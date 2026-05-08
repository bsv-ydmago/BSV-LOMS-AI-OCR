"""
lu_simulator_patch.py — Risk Simulator Tab
============================================
Inflation / Cost-Shock Simulator.

This file is the CANONICAL simulator implementation.
It supersedes _sim_* functions in lu_ui.py and the simulator patches
in lu_analysis_patch_v8.py.  Apply it AFTER lu_tab_analysis.attach(cls).

Standalone: imports only lu_core and lu_shared.
Attached to app class via attach(cls).

Key improvements over original lu_ui
--------------------------------------
  • Expense mix shown as a matplotlib pie (% of total simulated spend).
  • Chunked row construction via after() to keep UI responsive.
  • SIM_MAX_ROWS cap (50) to prevent widget explosion.
  • Respects active sector filter from lu_shared.
"""

import tkinter as tk
from tkinter import messagebox, filedialog
import tkinter.ttk as ttk
import customtkinter as ctk
import re

from lu_core import GENERAL_CLIENT, _RISK_ORDER, _fmt_value, get_high_risk_industries
from lu_shared import (
    F, FF, _bind_mousewheel,
    _NAVY_DEEP, _NAVY_MID, _NAVY_LIGHT, _NAVY_MIST, _NAVY_GHOST, _NAVY_PALE,
    _WHITE, _CARD_WHITE, _OFF_WHITE, _BORDER_LIGHT, _BORDER_MID,
    _TXT_NAVY, _TXT_SOFT, _TXT_MUTED, _TXT_ON_LIME,
    _LIME_MID, _LIME_DARK, _LIME_PALE,
    _ACCENT_RED, _ACCENT_GOLD, _ACCENT_SUCCESS,
    _RISK_COLOR, _RISK_BG, _RISK_BADGE_BG,
    _lu_filter_data_by_query,
    _lu_get_active_sectors, _lu_get_filtered_all_data,
    LU_CLIENT_TREE_SPEC, lu_format_lu_cell,
)

# ── Tuneable constants ──────────────────────────────────────────────
SIM_MAX_ROWS       = 50   # max expense rows shown
SIM_CHUNK_SIZE     = 10   # rows built per after() tick
SIM_CHART_MAX_BARS = 20   # max expense rows considered for chart
PIE_MAX_SLICES     = 10   # pie slices before aggregating to "Other"
SIM_TABLE_COLUMNS = (
    # (title, min_width_px, weight)
    ("Expense Item", 220, 5),
    ("Risk", 72, 1),
    ("Base Amount", 120, 2),
    ("Inflation Rate (%)", 100, 2),
    ("Extra Cost", 120, 2),
    ("Simulated", 120, 2),
)

SIM_CLIENT_TABLE_COLUMNS = (
    # (col_id, heading, min_width_px, anchor)
    ("client",               "Client Name",           210, "w"),
    ("industry",             "Industry",              150, "w"),
    ("base_total_expenses",  "Total Expenses (Base)", 150, "e"),
    ("sim_total_expenses", "Total Expenses (Sim)", 150, "e"),
    ("net_income", "Total Net Income (Base)", 160, "e"),
    ("sim_net_income", "Total Net Income (Simulated)", 185, "e"),
    ("pct_increase", "% Increase", 90, "center"),
    ("sim_increase", "Simulated Increase", 150, "e"),
    ("current_amort", "Total Current Amort", 150, "e"),
    ("pct_net_to_amort", "% Net → Amort", 120, "center"),
    ("sim_risk_label", "Risk Label", 90, "center"),
    ("risk_reasoning", "Risk Reasoning", 280, "w"),
)

SIM_CLIENT_PAGE_SIZE = 10

# ── Default risk range boundaries (% Net → Amort) ───────────────────
# Each entry: (min_inclusive, max_inclusive)
# Stored on the app instance as self._sim_risk_ranges at runtime.
_SIM_DEFAULT_RISK_RANGES = {
    "LOW":    (1.0,  35.0),
    "MEDIUM": (36.0, 70.0),
    "HIGH":   (70.01, float("inf")),
}

def _apply_sim_client_tree_style():
    """
    Match the Summary tab Treeview look-and-feel.
    (Same row height, heading padding, clam theme, and selection colors.)
    """
    style = ttk.Style()
    try:
        style.theme_use("clam")
    except Exception:
        pass
    style.configure(
        "SimSummary.Treeview",
        background=_WHITE,
        foreground=_TXT_NAVY,
        fieldbackground=_WHITE,
        rowheight=36,
        font=("Segoe UI", 9),
        borderwidth=0,
        relief="flat",
    )
    style.configure(
        "SimSummary.Treeview.Heading",
        background=_NAVY_DEEP,
        foreground=_WHITE,
        font=("Segoe UI", 9, "bold"),
        relief="flat",
        borderwidth=0,
        padding=(10, 8),
    )
    style.map(
        "SimSummary.Treeview.Heading",
        background=[("active", _NAVY_LIGHT)],
        relief=[("active", "flat")],
    )
    style.map(
        "SimSummary.Treeview",
        background=[("selected", "#C8E6C9")],
        foreground=[("selected", _NAVY_DEEP)],
    )

try:
    import matplotlib
    matplotlib.use("TkAgg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
    _HAS_MPL = True
except ImportError:
    _HAS_MPL = False


# ══════════════════════════════════════════════════════════════════════
#  PANEL BUILDER (sets fixed canvas height)
# ══════════════════════════════════════════════════════════════════════

def _build_simulator_panel(self, parent):
    hdr = tk.Frame(parent, bg=_NAVY_MID, height=38)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    self._sim_hdr_lbl = tk.Label(
        hdr, text="⚙️  Inflation / Cost-Shock Simulator",
        font=F(10, "bold"), fg=_WHITE, bg=_NAVY_MID)
    self._sim_hdr_lbl.pack(side="left", padx=20, pady=8)

    ctrl = tk.Frame(parent, bg=_OFF_WHITE, height=58)
    ctrl.pack(fill="x")
    ctrl.pack_propagate(False)
    # kept as attrs for compatibility with _sim_apply_global / _sim_reset / _sim_populate
    self._sim_global_var = tk.StringVar(value="0")
    self._sim_search_var = tk.StringVar()
    self._sim_match_lbl = tk.Label(
        ctrl, text="", font=F(8, "bold"), fg=_WHITE, bg=_OFF_WHITE, padx=8, pady=3)
    ctk.CTkButton(
        ctrl,
        text="⧉  Expense Simulator Table",
        command=lambda: _sim_open_expense_table_window(self),
        width=200,
        height=32,
        corner_radius=5,
        fg_color=_NAVY_LIGHT,
        hover_color=_NAVY_MID,
        text_color=_WHITE,
        font=FF(10, "bold"),
    ).pack(side="left", padx=(16, 0), pady=13)
    self._sim_industry_filter_btn = ctk.CTkButton(
        ctrl,
        text="Industry Checklist",
        command=lambda: _sim_open_industry_checklist(self),
        width=150,
        height=32,
        corner_radius=5,
        fg_color=_NAVY_LIGHT,
        hover_color=_NAVY_MID,
        text_color=_WHITE,
        font=FF(10, "bold"),
    )
    self._sim_industry_filter_btn.pack(side="left", padx=(8, 0), pady=10)
    self._sim_industry_filter_lbl = tk.Label(
        ctrl, text="", font=F(8, "bold"), fg=_TXT_SOFT, bg=_OFF_WHITE
    )
    self._sim_industry_filter_lbl.pack(side="left", padx=(8, 0), pady=10)
    ctk.CTkButton(
        ctrl,
        text="⚖  Risk Ranges",
        command=lambda: _sim_open_risk_ranges_dialog(self),
        width=130,
        height=32,
        corner_radius=5,
        fg_color=_NAVY_LIGHT,
        hover_color=_NAVY_MID,
        text_color=_WHITE,
        font=FF(10, "bold"),
    ).pack(side="left", padx=(8, 0), pady=10)
    tk.Frame(parent, bg=_BORDER_LIGHT, height=1).pack(fill="x")

    cards_frame = tk.Frame(parent, bg=_NAVY_MIST)
    cards_frame.pack(fill="x")
    _build_sim_summary_cards(self, cards_frame)

    # inc_bar removed — surplus/deficit label no longer displayed here.
    # Store dummy label attrs so _sim_refresh does not crash on hasattr checks.
    self._sim_income_lbl  = tk.Label(parent, text="")
    self._sim_surplus_lbl = tk.Label(parent, text="")

    # ── Outer scrollable body — one canvas scrolls everything ──────────
    body = tk.Frame(parent, bg=_CARD_WHITE)
    body.pack(fill="both", expand=True)

    _outer_vsb = tk.Scrollbar(body, orient="vertical", relief="flat",
                               troughcolor=_OFF_WHITE, bg=_BORDER_LIGHT, width=10, bd=0)
    _outer_vsb.pack(side="right", fill="y")
    _outer_canvas = tk.Canvas(body, bg=_CARD_WHITE, highlightthickness=0,
                               yscrollcommand=_outer_vsb.set)
    _outer_canvas.pack(side="left", fill="both", expand=True)
    _outer_vsb.config(command=_outer_canvas.yview)

    _outer_frame = tk.Frame(_outer_canvas, bg=_CARD_WHITE)
    _outer_win = _outer_canvas.create_window((0, 0), window=_outer_frame, anchor="nw")

    def _outer_on_frame_configure(e):
        _outer_canvas.configure(scrollregion=_outer_canvas.bbox("all"))
    _outer_frame.bind("<Configure>", _outer_on_frame_configure)

    def _outer_on_canvas_configure(e):
        _outer_canvas.itemconfig(_outer_win, width=e.width)
    _outer_canvas.bind("<Configure>", _outer_on_canvas_configure)

    def _outer_mousewheel(event):
        _outer_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
    _outer_canvas.bind("<Enter>", lambda _e: _outer_canvas.bind_all("<MouseWheel>", _outer_mousewheel))
    _outer_canvas.bind("<Leave>", lambda _e: _outer_canvas.unbind_all("<MouseWheel>"))

    # ── CLIENT IMPACT SECTION ─────────────────────────────────────────
    # Two stacked sub-frames inside a shared container.
    # Only one is visible at a time:
    #   _sim_client_table_view  — the treeview + pagination (default)
    #   _sim_client_detail_view — the full single-client panel with Back btn
    # ──────────────────────────────────────────────────────────────────
    _client_section = tk.Frame(_outer_frame, bg=_CARD_WHITE)
    _client_section.pack(fill="x", expand=False)

    # Section header bar removed. Back button is now rendered inside the detail
    # view content by _sim_show_client_details. Keep dummy attrs so show/hide
    # helpers do not crash.
    self._sim_section_title_lbl = tk.Label(parent, text="")
    self._sim_back_btn = None

    # ═══════════════════════════════════════════════════════════════════
    #  TABLE VIEW  (treeview + search + pagination)
    # ═══════════════════════════════════════════════════════════════════
    self._sim_client_table_view = tk.Frame(_client_section, bg=_CARD_WHITE)
    self._sim_client_table_view.pack(fill="x", expand=False)
    table_wrap = self._sim_client_table_view  # alias so existing code below works

    # Header row: label on left, search bar on right
    hdr_row = tk.Frame(table_wrap, bg=_CARD_WHITE)
    hdr_row.pack(fill="x", padx=20, pady=(10, 4))
    tk.Label(
        hdr_row,
        text="Client Impact (updates when you ramp expenses)",
        font=F(9, "bold"),
        fg=_TXT_SOFT,
        bg=_CARD_WHITE,
    ).pack(side="left")
    self._sim_client_search_var = tk.StringVar()
    _search_frame = tk.Frame(hdr_row, bg=_CARD_WHITE)
    _search_frame.pack(side="right")
    tk.Label(
        _search_frame,
        text="🔍",
        font=F(10),
        fg=_TXT_MUTED,
        bg=_CARD_WHITE,
    ).pack(side="left", padx=(0, 4))
    _client_search_entry = ctk.CTkEntry(
        _search_frame,
        textvariable=self._sim_client_search_var,
        width=220,
        height=28,
        corner_radius=6,
        fg_color=_WHITE,
        text_color=_TXT_NAVY,
        border_color=_BORDER_MID,
        font=FF(10),
        placeholder_text="Search client name…",
    )
    _client_search_entry.pack(side="left")
    def _client_search_refresh():
        self._sim_client_page = 0
        _sim_populate(self)
    _client_search_entry.bind("<KeyRelease>", lambda _e: _client_search_refresh())
    ctk.CTkButton(
        _search_frame,
        text="✕",
        width=28,
        height=28,
        corner_radius=6,
        fg_color=_OFF_WHITE,
        hover_color=_BORDER_MID,
        text_color=_TXT_MUTED,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: [
            self._sim_client_search_var.set(""),
            _client_search_refresh(),
        ],
    ).pack(side="left", padx=(4, 0))

    tk.Label(
        table_wrap,
        text="Risk Label is based on % Net Income to Amortization. Use ⚖ Risk Ranges to configure thresholds.",
        font=F(7),
        fg=_TXT_MUTED,
        bg=_CARD_WHITE,
    ).pack(anchor="w", padx=20, pady=(0, 6))

    # Page controls
    pg_row = tk.Frame(table_wrap, bg=_CARD_WHITE)
    pg_row.pack(fill="x", padx=20, pady=(0, 6))
    self._sim_client_prev_btn = ctk.CTkButton(
        pg_row,
        text="◀ Prev",
        width=70,
        height=26,
        corner_radius=6,
        fg_color=_WHITE,
        hover_color=_NAVY_MIST,
        text_color=_NAVY_MID,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: _sim_client_page_prev(self),
        state="disabled",
    )
    self._sim_client_prev_btn.pack(side="left")
    self._sim_client_page_lbl = tk.Label(
        pg_row,
        text="Page 1",
        font=F(8),
        fg=_TXT_SOFT,
        bg=_CARD_WHITE,
    )
    self._sim_client_page_lbl.pack(side="left", padx=10)
    self._sim_client_next_btn = ctk.CTkButton(
        pg_row,
        text="Next ▶",
        width=70,
        height=26,
        corner_radius=6,
        fg_color=_WHITE,
        hover_color=_NAVY_MIST,
        text_color=_NAVY_MID,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: _sim_client_page_next(self),
        state="disabled",
    )
    self._sim_client_next_btn.pack(side="left")

    self._sim_export_btn = ctk.CTkButton(
        pg_row,
        text="💾  Simulator Clients Excel",
        width=100,
        height=26,
        corner_radius=6,
        fg_color=_WHITE,
        hover_color=_NAVY_MIST,
        text_color=_NAVY_MID,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: _sim_export_client_impact_excel(self),
    )
    self._sim_export_btn.pack(side="right", padx=(0, 10))
    self._sim_export_high_risk_btn = ctk.CTkButton(
        pg_row,
        text="🔴  HIGH Risk Clients Excel",
        width=100,
        height=26,
        corner_radius=6,
        fg_color="#FFF0F0",
        hover_color="#FFD6D6",
        text_color=_ACCENT_RED,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_ACCENT_RED,
        command=lambda: _sim_export_high_risk_clients_excel(self),
    )
    self._sim_export_high_risk_btn.pack(side="right", padx=(0, 6))
    self._sim_merge_excel_btn = ctk.CTkButton(
        pg_row,
        text="🧩  Merge Excel",
        width=92,
        height=26,
        corner_radius=6,
        fg_color="#EEF3FA",
        hover_color="#DCE7F6",
        text_color=_NAVY_MID,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: _sim_merge_excel_files(self),
    )
    self._sim_merge_excel_btn.pack(side="right", padx=(0, 6))
    self._sim_client_count_lbl = tk.Label(
        pg_row,
        text="",
        font=F(8),
        fg=_TXT_MUTED,
        bg=_CARD_WHITE,
    )
    self._sim_client_count_lbl.pack(side="right")

    # Treeview
    tbl_box = tk.Frame(table_wrap, bg=_CARD_WHITE)
    tbl_box.pack(fill="x", padx=20, pady=(0, 16))

    hsb = ttk.Scrollbar(tbl_box, orient="horizontal")
    hsb.pack(side="bottom", fill="x")

    _apply_sim_client_tree_style()
    self._sim_client_tree = ttk.Treeview(
        tbl_box,
        columns=tuple(c[0] for c in SIM_CLIENT_TABLE_COLUMNS),
        show="headings",
        height=10,
        style="SimSummary.Treeview",
        xscrollcommand=hsb.set,
        selectmode="browse",
    )
    self._sim_client_tree.pack(side="top", fill="x")
    hsb.config(command=self._sim_client_tree.xview)

    self._sim_client_tree.bind(
        "<ButtonRelease-1>",
        lambda e: _sim_on_client_impact_row_activated(self, e),
    )
    self._sim_client_tree.bind(
        "<Return>",
        lambda e: _sim_on_client_impact_row_activated(self, e),
    )

    for col_id, heading, min_px, anchor in SIM_CLIENT_TABLE_COLUMNS:
        self._sim_client_tree.heading(col_id, text=heading)
        stretch = (col_id == "client")
        self._sim_client_tree.column(col_id, width=min_px, minwidth=min_px, anchor=anchor, stretch=stretch)

    # Risk tags
    self._sim_client_tree.tag_configure("HIGH",   background="#FFF5F5", foreground=_ACCENT_RED)
    self._sim_client_tree.tag_configure("HIGH_NEG", background="#FFE5E5", foreground=_ACCENT_RED)
    self._sim_client_tree.tag_configure("MEDIUM", background="#FFFBF0", foreground=_ACCENT_GOLD)
    self._sim_client_tree.tag_configure("LOW",    background="#F0FBE8", foreground=_ACCENT_SUCCESS)
    self._sim_client_tree.tag_configure("NA",     background=_WHITE,    foreground=_TXT_MUTED)

    # Hover tag — blue highlight matching Analysis tab
    self._sim_client_tree.tag_configure(
        "row_hover",
        background="#D6E8FF",
        foreground=_NAVY_DEEP,
    )

    # Hand cursor
    self._sim_client_tree.configure(cursor="hand2")

    # Hover motion bindings
    _sim_tree_hovered_iid = [None]
    self._sim_client_tree._sim_hovered_iid_ref = _sim_tree_hovered_iid

    def _sim_tree_on_motion(event, _tree=self._sim_client_tree,
                            _hovered=_sim_tree_hovered_iid):
        iid = _tree.identify_row(event.y)
        prev = _hovered[0]
        if iid == prev:
            return
        if prev and prev in _tree.get_children(""):
            tags = [t for t in _tree.item(prev, "tags") if t != "row_hover"]
            _tree.item(prev, tags=tags)
        if iid and iid in _tree.get_children(""):
            tags = [t for t in _tree.item(iid, "tags") if t != "row_hover"]
            tags.append("row_hover")
            _tree.item(iid, tags=tags)
        _hovered[0] = iid

    def _sim_tree_on_leave(event, _tree=self._sim_client_tree,
                           _hovered=_sim_tree_hovered_iid):
        prev = _hovered[0]
        if prev and prev in _tree.get_children(""):
            tags = [t for t in _tree.item(prev, "tags") if t != "row_hover"]
            _tree.item(prev, tags=tags)
        _hovered[0] = None

    self._sim_client_tree.bind("<Motion>", _sim_tree_on_motion)
    self._sim_client_tree.bind("<Leave>",  _sim_tree_on_leave)

    # Hint strip at bottom of table view
    _tbl_hint = tk.Frame(table_wrap, bg=_NAVY_MIST,
                         highlightbackground=_BORDER_MID, highlightthickness=1)
    _tbl_hint.pack(fill="x", padx=20, pady=(0, 16))
    tk.Label(
        _tbl_hint,
        text="👆  Click any row to view the full client details.",
        font=F(8),
        fg=_TXT_SOFT,
        bg=_NAVY_MIST,
        anchor="w",
        padx=14,
        pady=8,
    ).pack(anchor="w")

    # ═══════════════════════════════════════════════════════════════════
    #  DETAIL VIEW  (single client — hidden until a row is clicked)
    # ═══════════════════════════════════════════════════════════════════
    self._sim_client_detail_view = tk.Frame(_client_section, bg=_CARD_WHITE)
    # NOT packed yet — shown on click, hidden on Back
    self._sim_client_detail_frame = self._sim_client_detail_view   # alias for _sim_show_client_details

    # Bottom padding spacer
    tk.Frame(_outer_frame, bg=_CARD_WHITE, height=20).pack(fill="x")

    self._sim_sliders    = {}
    self._sim_expenses   = []
    self._sim_build_job  = None
    self._sim_expenses_capped = False
    self._sim_expense_win = None
    self._sim_expense_search_var = tk.StringVar()
    self._sim_canvas = None
    self._sim_scroll_frame = None
    self._sim_net_income = 0.0
    self._sim_recs = []
    self._sim_client_page = 0
    self._sim_selected_industries = set()
    self._sim_manual_industries = set()
    self._sim_chart_holder = None  # chart removed; kept for compatibility
    _sim_show_placeholder(self)
def _sim_filter_data_by_industry_checklist(all_data, selected_industries):
    selected = {str(x).strip().lower() for x in (selected_industries or set()) if str(x).strip()}
    if not selected:
        return all_data

    splitter = re.compile(r"\s*(?:,|/|;|&|\band\b)\s*", re.I)

    def _industry_tokens(rec: dict) -> set[str]:
        tags = rec.get("industry_tags") or []
        if tags:
            return {str(x).strip().lower() for x in tags if str(x).strip()}
        raw = str((rec or {}).get("industry") or "").strip()
        if not raw:
            return set()
        return {tok.strip().lower() for tok in splitter.split(raw) if tok.strip()}

    base_general = list((all_data or {}).get("general", []))
    kept_general = [
        rec for rec in base_general
        if _industry_tokens(rec) & selected
    ]
    kept_clients = {str((r or {}).get("client") or ""): r for r in kept_general if (r or {}).get("client")}

    patched = dict(all_data or {})
    patched["general"] = kept_general
    patched["clients"] = kept_clients
    return patched


def _sim_open_industry_checklist(self):
    all_data = getattr(self, "_lu_all_data", None) or {}
    base_industries = {str(x).strip() for x in all_data.get("unique_industries", []) if str(x).strip()}
    manual_industries = set(getattr(self, "_sim_manual_industries", set()) or set())
    industries = sorted(base_industries | manual_industries, key=str.lower)
    if not industries:
        messagebox.showwarning("No Data", "Load and run LU analysis first.")
        return

    selected = set(getattr(self, "_sim_selected_industries", set()) or set())
    high_defaults = {str(x).strip().lower() for x in get_high_risk_industries() if str(x).strip()}
    if not selected:
        selected = {name for name in industries if name.lower() in high_defaults}

    dialog = ctk.CTkToplevel(self)
    dialog.title("Risk Simulator Industry Checklist")
    dialog.geometry("620x620")
    dialog.minsize(520, 480)
    dialog.transient(self)
    dialog.grab_set()
    dialog.configure(fg_color=_CARD_WHITE)

    hdr = tk.Frame(dialog, bg=_NAVY_MID, height=52)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    tk.Label(
        hdr,
        text="☑  Risk Simulator Industry Checklist",
        font=F(11, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
    ).pack(side="left", padx=16, pady=12)

    info = tk.Frame(dialog, bg=_NAVY_MIST, highlightbackground=_BORDER_MID, highlightthickness=1)
    info.pack(fill="x", padx=16, pady=(10, 6))
    tk.Label(
        info,
        text=(
            "Use checkboxes to filter the Risk Simulator by industry. "
            "By default, this follows HIGH industries from Risk Settings."
        ),
        font=F(8),
        fg=_TXT_SOFT,
        bg=_NAVY_MIST,
        anchor="w",
        justify="left",
    ).pack(fill="x", padx=10, pady=8)

    search_row = tk.Frame(dialog, bg=_CARD_WHITE)
    search_row.pack(fill="x", padx=16, pady=(4, 6))
    tk.Label(search_row, text="🔍", font=F(10), fg=_TXT_SOFT, bg=_CARD_WHITE).pack(side="left")
    search_var = tk.StringVar(value="")
    search_entry = ctk.CTkEntry(
        search_row,
        textvariable=search_var,
        width=380,
        height=28,
        corner_radius=4,
        fg_color=_WHITE,
        text_color=_TXT_NAVY,
        border_color=_BORDER_MID,
        font=FF(9),
        placeholder_text="Search industry...",
    )
    search_entry.pack(side="left", fill="x", expand=True, padx=(6, 0))

    add_row = tk.Frame(dialog, bg=_CARD_WHITE)
    add_row.pack(fill="x", padx=16, pady=(0, 6))
    tk.Label(
        add_row, text="Add Industry:", font=F(8, "bold"),
        fg=_NAVY_MID, bg=_CARD_WHITE
    ).pack(side="left", padx=(0, 6))
    add_var = tk.StringVar(value="")
    add_entry = ctk.CTkEntry(
        add_row,
        textvariable=add_var,
        width=260,
        height=28,
        corner_radius=4,
        fg_color=_WHITE,
        text_color=_TXT_NAVY,
        border_color=_BORDER_MID,
        font=FF(9),
        placeholder_text="e.g. Logistics",
    )
    add_entry.pack(side="left", padx=(0, 6))
    add_hint_lbl = tk.Label(add_row, text="", font=F(7), fg=_TXT_MUTED, bg=_CARD_WHITE)
    add_hint_lbl.pack(side="left", padx=(4, 0))

    list_wrap = tk.Frame(dialog, bg=_CARD_WHITE)
    list_wrap.pack(fill="both", expand=True, padx=16, pady=(0, 8))
    sb = tk.Scrollbar(list_wrap, relief="flat", troughcolor=_OFF_WHITE, bg=_BORDER_LIGHT, width=8, bd=0)
    sb.pack(side="right", fill="y")
    canvas = tk.Canvas(list_wrap, bg=_CARD_WHITE, highlightthickness=0, yscrollcommand=sb.set)
    canvas.pack(side="left", fill="both", expand=True)
    sb.config(command=canvas.yview)
    rows_frame = tk.Frame(canvas, bg=_CARD_WHITE)
    win = canvas.create_window((0, 0), window=rows_frame, anchor="nw")
    rows_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(win, width=e.width))
    canvas.bind("<Enter>", lambda _e: canvas.bind_all("<MouseWheel>", lambda ev: canvas.yview_scroll(int(-1 * (ev.delta / 120)), "units")))
    canvas.bind("<Leave>", lambda _e: canvas.unbind_all("<MouseWheel>"))

    row_widgets = []
    var_map = {}
    row_by_industry = {}

    def _add_industry_row(industry: str, preselect: bool = False):
        if industry in var_map:
            if preselect:
                var_map[industry].set(True)
            return
        idx = len(row_widgets)
        row_bg = _WHITE if idx % 2 == 0 else _OFF_WHITE
        row = tk.Frame(rows_frame, bg=row_bg)
        row.pack(fill="x")
        tk.Frame(row, bg=_BORDER_LIGHT, height=1).pack(fill="x")
        inner = tk.Frame(row, bg=row_bg)
        inner.pack(fill="x", padx=8, pady=4)
        var = tk.BooleanVar(value=preselect)
        var_map[industry] = var
        chk = tk.Checkbutton(
            inner,
            text=industry,
            variable=var,
            onvalue=True,
            offvalue=False,
            font=F(9),
            fg=_TXT_NAVY,
            bg=row_bg,
            activebackground=row_bg,
            anchor="w",
            justify="left",
            relief="flat",
            highlightthickness=0,
        )
        chk.pack(side="left", fill="x", expand=True, padx=(4, 0))
        row_widgets.append((row, industry))
        row_by_industry[industry] = row

    for industry in industries:
        _add_industry_row(industry, preselect=(industry in selected))

    def _apply_search(*_args):
        q = search_var.get().strip().lower()
        for row, industry in row_widgets:
            show = (not q) or (q in industry.lower())
            if show and not row.winfo_ismapped():
                row.pack(fill="x")
            elif (not show) and row.winfo_ismapped():
                row.pack_forget()

    search_var.trace_add("write", _apply_search)

    def _add_manual_industry():
        name = add_var.get().strip()
        if not name:
            add_hint_lbl.config(text="Enter an industry name.", fg=_ACCENT_RED)
            return
        existing = {k.lower(): k for k in var_map.keys()}
        if name.lower() in existing:
            _add_industry_row(existing[name.lower()], preselect=True)
            add_hint_lbl.config(text="Already exists; selected.", fg=_ACCENT_SUCCESS)
        else:
            _add_industry_row(name, preselect=True)
            add_hint_lbl.config(text=f"Added '{name}'.", fg=_ACCENT_SUCCESS)
        add_var.set("")
        _apply_search()
        add_entry.focus_set()

    def _remove_manual_industry():
        name = add_var.get().strip()
        if not name:
            add_hint_lbl.config(text="Enter an industry to remove.", fg=_ACCENT_RED)
            return
        existing = {k.lower(): k for k in var_map.keys()}
        canonical = existing.get(name.lower())
        if not canonical:
            add_hint_lbl.config(text="Industry not found.", fg=_ACCENT_RED)
            return
        if canonical in base_industries:
            add_hint_lbl.config(text="Cannot remove base industry from data.", fg=_ACCENT_RED)
            return
        row = row_by_industry.pop(canonical, None)
        if row is not None:
            try:
                row.destroy()
            except Exception:
                pass
        row_widgets[:] = [(r, n) for (r, n) in row_widgets if n != canonical]
        var_map.pop(canonical, None)
        selected.discard(canonical)
        add_hint_lbl.config(text=f"Removed '{canonical}'.", fg=_ACCENT_SUCCESS)
        add_var.set("")
        _apply_search()
        add_entry.focus_set()

    tk.Button(
        add_row,
        text="Add",
        font=F(8, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
        activebackground=_NAVY_LIGHT,
        activeforeground=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=5,
        cursor="hand2",
        command=_add_manual_industry,
    ).pack(side="left")
    tk.Button(
        add_row,
        text="Remove",
        font=F(8, "bold"),
        fg=_WHITE,
        bg=_ACCENT_RED,
        activebackground="#C53030",
        activeforeground=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=5,
        cursor="hand2",
        command=_remove_manual_industry,
    ).pack(side="left", padx=(4, 0))
    add_entry.bind("<Return>", lambda _e: _add_manual_industry())

    def _set_all(v: bool):
        for _name, vv in var_map.items():
            vv.set(v)

    def _use_high_defaults():
        for name, vv in var_map.items():
            vv.set(name.lower() in high_defaults)

    def _apply_and_close():
        chosen = {name for name, vv in var_map.items() if vv.get()}
        self._sim_selected_industries = chosen
        self._sim_manual_industries = set(var_map.keys()) - base_industries
        _sim_populate(self)
        dialog.destroy()

    footer = tk.Frame(dialog, bg=_OFF_WHITE, highlightbackground=_BORDER_MID, highlightthickness=1)
    footer.pack(fill="x", padx=16, pady=(2, 14))
    tk.Button(
        footer,
        text="Select All",
        font=F(8, "bold"),
        fg=_TXT_NAVY,
        bg=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=6,
        cursor="hand2",
        command=lambda: _set_all(True),
    ).pack(side="left", padx=(12, 4), pady=8)
    tk.Button(
        footer,
        text="Clear All",
        font=F(8, "bold"),
        fg=_TXT_SOFT,
        bg=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=6,
        cursor="hand2",
        command=lambda: _set_all(False),
    ).pack(side="left", padx=4, pady=8)
    tk.Button(
        footer,
        text="Use HIGH from Settings",
        font=F(8, "bold"),
        fg=_ACCENT_RED,
        bg="#FFE8E8",
        relief="flat",
        bd=0,
        padx=10,
        pady=6,
        cursor="hand2",
        command=_use_high_defaults,
    ).pack(side="left", padx=4, pady=8)
    tk.Button(
        footer,
        text="Cancel",
        font=F(9),
        fg=_TXT_SOFT,
        bg=_OFF_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=8,
        cursor="hand2",
        command=dialog.destroy,
    ).pack(side="right", padx=(0, 4), pady=8)
    tk.Button(
        footer,
        text="  ✔  Apply Filter  ",
        font=F(9, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
        activebackground=_NAVY_LIGHT,
        activeforeground=_WHITE,
        relief="flat",
        bd=0,
        padx=14,
        pady=8,
        cursor="hand2",
        command=_apply_and_close,
    ).pack(side="right", padx=12, pady=8)






def _sim_open_risk_ranges_dialog(self):
    """
    Settings dialog for configuring custom risk range boundaries.
    Mirrors the style of the Industry Checklist dialog.

    Stores result on self._sim_risk_ranges as:
        {"LOW": (min, max), "MEDIUM": (min, max), "HIGH": (min, max)}
    where HIGH max is float('inf').

    Validates that no two ranges share any overlapping percentage value.
    """
    # Load current ranges (or defaults)
    current = dict(getattr(self, "_sim_risk_ranges", None) or _SIM_DEFAULT_RISK_RANGES)

    dialog = ctk.CTkToplevel(self)
    dialog.title("Risk Range Settings")
    dialog.geometry("500x440")
    dialog.minsize(460, 400)
    dialog.resizable(False, False)
    dialog.transient(self)
    dialog.grab_set()
    dialog.configure(fg_color=_CARD_WHITE)

    # ── Header ────────────────────────────────────────────────────────
    hdr = tk.Frame(dialog, bg=_NAVY_MID, height=52)
    hdr.pack(fill="x")
    hdr.pack_propagate(False)
    tk.Label(
        hdr,
        text="⚖  Risk Range Settings",
        font=F(11, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
    ).pack(side="left", padx=16, pady=12)

    # ── Info banner ───────────────────────────────────────────────────
    info = tk.Frame(dialog, bg=_NAVY_MIST, highlightbackground=_BORDER_MID, highlightthickness=1)
    info.pack(fill="x", padx=16, pady=(12, 6))
    tk.Label(
        info,
        text=(
            "Set the % Net → Amort boundaries for each risk level.\n"
            "HIGH max is always open-ended (any value above HIGH min is HIGH).\n"
            "Ranges must not overlap — each percentage point belongs to exactly one level."
        ),
        font=F(8),
        fg=_TXT_SOFT,
        bg=_NAVY_MIST,
        anchor="w",
        justify="left",
        wraplength=440,
    ).pack(fill="x", padx=10, pady=8)

    # ── Warning label (shown when overlap detected) ────────────────────
    warn_var = tk.StringVar(value="")
    warn_lbl = tk.Label(
        dialog,
        textvariable=warn_var,
        font=F(8, "bold"),
        fg=_ACCENT_RED,
        bg="#FFF5F5",
        anchor="w",
        justify="left",
        wraplength=440,
        padx=10,
        pady=6,
        relief="flat",
    )
    # packed only when there is a warning — we pack/forget it dynamically

    # ── Range input rows ──────────────────────────────────────────────
    RISKS = [
        ("LOW",    _ACCENT_SUCCESS, "#F0FBE8"),
        ("MEDIUM", _ACCENT_GOLD,   "#FFFBF0"),
        ("HIGH",   _ACCENT_RED,    "#FFF5F5"),
    ]

    fields: dict[str, dict] = {}   # "LOW" / "MEDIUM" / "HIGH" → {"min": var, "max": var}

    rows_frame = tk.Frame(dialog, bg=_CARD_WHITE)
    rows_frame.pack(fill="x", padx=16, pady=(4, 0))

    for risk, badge_color, row_bg in RISKS:
        lo, hi = current.get(risk, _SIM_DEFAULT_RISK_RANGES[risk])
        hi_display = "" if hi == float("inf") else str(int(hi) if hi == int(hi) else hi)
        lo_display = str(int(lo) if lo == int(lo) else lo)

        row = tk.Frame(rows_frame, bg=row_bg,
                       highlightbackground=_BORDER_MID, highlightthickness=1)
        row.pack(fill="x", pady=4)

        # Badge strip (colored left border)
        badge = tk.Frame(row, bg=badge_color, width=12)
        badge.pack(side="left", fill="y")
        badge.pack_propagate(False)

        # Risk label as colored text
        tk.Label(
            row,
            text=risk,
            font=F(10, "bold"),
            fg=badge_color,
            bg=row_bg,
            width=8,
            anchor="w",
            padx=10,
        ).pack(side="left")

        # Min field
        tk.Label(row, text="Min %", font=F(8), fg=_TXT_SOFT, bg=row_bg, padx=(6)).pack(side="left")
        min_var = tk.StringVar(value=lo_display)
        min_entry = ctk.CTkEntry(
            row,
            textvariable=min_var,
            width=72,
            height=30,
            corner_radius=4,
            fg_color=_WHITE,
            text_color=_TXT_NAVY,
            border_color=badge_color,
            font=FF(10),
        )
        min_entry.pack(side="left", padx=(4, 8))

        # Max field
        if risk == "HIGH":
            tk.Label(row, text="Max %  ∞ (open-ended)", font=F(8), fg=_TXT_MUTED, bg=row_bg, padx=6).pack(side="left")
            max_var = tk.StringVar(value="")
            max_entry = None
        else:
            tk.Label(row, text="Max %", font=F(8), fg=_TXT_SOFT, bg=row_bg, padx=6).pack(side="left")
            max_var = tk.StringVar(value=hi_display)
            max_entry = ctk.CTkEntry(
                row,
                textvariable=max_var,
                width=72,
                height=30,
                corner_radius=4,
                fg_color=_WHITE,
                text_color=_TXT_NAVY,
                border_color=badge_color,
                font=FF(10),
            )
            max_entry.pack(side="left", padx=(4, 8))

        fields[risk] = {"min": min_var, "max": max_var, "min_entry": min_entry,
                        "max_entry": max_entry, "bg": row_bg, "color": badge_color}

    # ── Live validation ───────────────────────────────────────────────
    def _parse_field(var: tk.StringVar, allow_empty: bool = False):
        """Return float or None on parse error. Empty string → None if allow_empty."""
        s = var.get().strip()
        if not s and allow_empty:
            return None
        try:
            return float(s)
        except ValueError:
            return None

    def _validate(*_args):
        """
        Parse all fields, detect overlaps, and update warn_lbl.
        Returns (ranges_dict, errors_list) — errors_list is [] if valid.

        Overlap rule: two inclusive ranges [lo1, hi1] and [lo2, hi2] overlap
        only when a value can belong to BOTH simultaneously.  Adjacent ranges
        like LOW(1–35) / MEDIUM(36–70) share no common value, so they are
        NOT treated as overlapping.  The test is:
            lo1 < hi2  AND  lo2 < hi1   (strict on both sides)
        For the HIGH range hi is ∞, so HIGH always overlaps anything whose
        lo < ∞ — but HIGH's lo is checked against the other ranges' hi values.
        """
        errors = []
        parsed: dict[str, tuple] = {}

        for risk, fd in fields.items():
            lo = _parse_field(fd["min"])
            if lo is None:
                errors.append(f"{risk}: Min % must be a number.")
                continue
            if risk == "HIGH":
                hi = float("inf")
            else:
                hi = _parse_field(fd["max"])
                if hi is None:
                    errors.append(f"{risk}: Max % must be a number.")
                    continue
                if hi < lo:
                    errors.append(f"{risk}: Max % must be ≥ Min %.")
                    continue
            parsed[risk] = (lo, hi)

        if not errors and len(parsed) == 3:
            # Skip overlap check if any required field is still empty (user is mid-edit).
            # HIGH's max is intentionally empty (open-ended), so skip it in this check.
            all_filled = all(
                fd["min"].get().strip() and (
                    risk == "HIGH" or fd["max"] is None or fd["max"].get().strip()
                )
                for risk, fd in fields.items()
            )
            if not all_filled:
                return None, []

            # Overlap check: two ranges share a value only when lo1 < hi2 AND lo2 < hi1
            # (strict inequalities so adjacent boundaries like 35 / 36 are allowed).
            overlap_msgs = []
            range_items = list(parsed.items())
            for i in range(len(range_items)):
                for j in range(i + 1, len(range_items)):
                    r1, (lo1, hi1) = range_items[i]
                    r2, (lo2, hi2) = range_items[j]
                    # Use strict < so adjacent non-overlapping ranges pass.
                    if lo1 < hi2 and lo2 < hi1:
                        def _fmt_r(lo, hi):
                            return f"{lo:.0f}%–{'∞' if hi == float('inf') else f'{hi:.0f}%'}"
                        msg = f"Overlap: {r1} ({_fmt_r(lo1, hi1)}) and {r2} ({_fmt_r(lo2, hi2)}) share values — fix the boundaries so no percentage belongs to two levels."
                        if msg not in overlap_msgs:
                            overlap_msgs.append(msg)
            if overlap_msgs:
                errors.extend(overlap_msgs)

        if errors:
            warn_var.set("⚠  " + "  |  ".join(errors))
            warn_lbl.pack(fill="x", padx=16, pady=(2, 4), before=rows_frame)
            return None, errors
        else:
            warn_var.set("")
            try:
                warn_lbl.pack_forget()
            except Exception:
                pass
            return parsed, []

    # Debounced live validation — waits 600 ms after last keystroke before running.
    # This prevents flooding the warning label while the user is mid-edit (e.g. backspacing).
    _debounce_job = [None]
    def _validate_debounced(*_args):
        if _debounce_job[0] is not None:
            try:
                dialog.after_cancel(_debounce_job[0])
            except Exception:
                pass
        _debounce_job[0] = dialog.after(600, _validate)
    for fd in fields.values():
        fd["min"].trace_add("write", _validate_debounced)
        if fd["max"] is not None:
            fd["max"].trace_add("write", _validate_debounced)

    # ── Reset to defaults ─────────────────────────────────────────────
    def _reset_defaults():
        for risk, fd in fields.items():
            lo, hi = _SIM_DEFAULT_RISK_RANGES[risk]
            fd["min"].set(str(int(lo) if lo == int(lo) else lo))
            if fd["max"] is not None:
                fd["max"].set(str(int(hi) if hi == int(hi) else hi))
        warn_var.set("")
        try:
            warn_lbl.pack_forget()
        except Exception:
            pass

    # ── Apply ─────────────────────────────────────────────────────────
    def _apply_and_close():
        # Cancel any pending debounce so we validate with the latest values.
        if _debounce_job[0] is not None:
            try:
                dialog.after_cancel(_debounce_job[0])
            except Exception:
                pass
            _debounce_job[0] = None

        parsed, errors = _validate()
        if errors:
            # Warning already shown by _validate — keep dialog open.
            return
        if parsed is None:
            warn_var.set("⚠  Please fill in all fields before applying.")
            warn_lbl.pack(fill="x", padx=16, pady=(2, 4), before=rows_frame)
            return

        self._sim_risk_ranges = parsed
        _sim_populate(self)
        dialog.destroy()

    # ── Footer ────────────────────────────────────────────────────────
    footer = tk.Frame(dialog, bg=_OFF_WHITE, highlightbackground=_BORDER_MID, highlightthickness=1)
    footer.pack(fill="x", padx=16, pady=(10, 14), side="bottom")

    tk.Button(
        footer,
        text="Reset to Defaults",
        font=F(8, "bold"),
        fg=_TXT_SOFT,
        bg=_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=6,
        cursor="hand2",
        command=_reset_defaults,
    ).pack(side="left", padx=(12, 4), pady=8)

    tk.Button(
        footer,
        text="Cancel",
        font=F(9),
        fg=_TXT_SOFT,
        bg=_OFF_WHITE,
        relief="flat",
        bd=0,
        padx=10,
        pady=8,
        cursor="hand2",
        command=dialog.destroy,
    ).pack(side="right", padx=(0, 4), pady=8)

    tk.Button(
        footer,
        text="  ✔  Apply Ranges  ",
        font=F(9, "bold"),
        fg=_WHITE,
        bg=_NAVY_MID,
        activebackground=_NAVY_LIGHT,
        activeforeground=_WHITE,
        relief="flat",
        bd=0,
        padx=14,
        pady=8,
        cursor="hand2",
        command=_apply_and_close,
    ).pack(side="right", padx=12, pady=8)

    # Run initial validation silently — only show warning if the pre-filled
    # values are genuinely invalid (not just because the user hasn't typed yet).
    _validate_debounced()


def _build_sim_summary_cards(self, parent):
    for title, attr, color in [
        ("Total Net Income",       "_sim_lbl_income",  _ACCENT_SUCCESS),
        ("Base Total Expenses",    "_sim_lbl_base",    _TXT_NAVY),
        ("Simulated Total",        "_sim_lbl_sim",     _TXT_NAVY),
        ("Total Increase (₱)",     "_sim_lbl_inc",     _ACCENT_RED),
        ("Surplus / Deficit",      "_sim_lbl_surplus", _ACCENT_SUCCESS),
    ]:
        card = tk.Frame(parent, bg=_NAVY_MIST,
                        highlightbackground="#D6E4F7", highlightthickness=1)
        card.pack(side="left", fill="x", expand=True, padx=6, pady=8)
        tk.Label(card, text=title, font=F(7), fg=_TXT_SOFT, bg=_NAVY_MIST
                 ).pack(anchor="w", padx=10, pady=(6, 0))
        lbl = tk.Label(card, text="—", font=F(13, "bold"), fg=color, bg=_NAVY_MIST)
        lbl.pack(anchor="w", padx=10, pady=(0, 6))
        setattr(self, attr, lbl)


# ══════════════════════════════════════════════════════════════════════
#  PLACEHOLDER
# ══════════════════════════════════════════════════════════════════════

def _sim_show_placeholder(self):
    frame = getattr(self, "_sim_scroll_frame", None)
    if frame is not None:
        for w in frame.winfo_children():
            w.destroy()
        tk.Label(
            frame,
            text="Run an analysis first to unlock the simulator.",
            font=F(10),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
        ).pack(pady=60)
    _sim_draw_chart(self)
    _sim_refresh_client_table(self)

    # Always return to table view and clear any client detail content.
    _sim_show_table_view(self)
    detail = getattr(self, "_sim_client_detail_frame", None)
    if detail is not None:
        for w in detail.winfo_children():
            try:
                w.destroy()
            except Exception:
                pass


# ══════════════════════════════════════════════════════════════════════
#  EXPENSE POPUP + POPULATE
# ══════════════════════════════════════════════════════════════════════

def _sim_render_expense_table_rows(self):
    """Render expense rows into the popup table (if open)."""
    frame = getattr(self, "_sim_scroll_frame", None)
    if frame is None:
        return
    try:
        if not frame.winfo_exists():
            return
    except Exception:
        return

    for w in list(frame.winfo_children()):
        try:
            w.destroy()
        except Exception:
            pass

    all_expenses = list(getattr(self, "_sim_expenses", []) or [])
    if not all_expenses:
        tk.Label(
            frame,
            text="No numeric expense data found.",
            font=F(9),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
            justify="center",
        ).pack(pady=60)
        return

    # Apply search filter if present
    _exp_q = str(getattr(self, "_sim_expense_search_var", None) and
                 self._sim_expense_search_var.get() or "").strip().lower()
    if _exp_q:
        all_expenses = [e for e in all_expenses if _exp_q in str(e.get("name") or "").lower()]

    if getattr(self, "_sim_expenses_capped", False):
        tk.Label(
            frame,
            text=f"ℹ  Showing top {SIM_MAX_ROWS} expense rows (file has more).",
            font=F(8),
            fg=_ACCENT_GOLD,
            bg=_OFF_WHITE,
            padx=10,
            pady=4,
        ).pack(fill="x")

    hdr = tk.Frame(frame, bg=_OFF_WHITE)
    hdr.pack(fill="x", pady=(8, 0))
    for col, (_title, min_px, _wt) in enumerate(SIM_TABLE_COLUMNS):
        hdr.grid_columnconfigure(col, weight=1, minsize=min_px, uniform="sim_col")
    for col, (text, _min_px, _wt) in enumerate(SIM_TABLE_COLUMNS):
        tk.Label(
            hdr,
            text=text,
            font=F(8, "bold"),
            fg=_NAVY_PALE,
            bg=_OFF_WHITE,
            anchor="w" if col == 0 else "center",
            justify="left" if col == 0 else "center",
            padx=6,
            pady=5,
        ).grid(row=0, column=col, sticky="ew", padx=(0, 2))
    tk.Frame(frame, bg=_BORDER_MID, height=1).pack(fill="x")

    for idx, exp in enumerate(all_expenses):
        var = self._sim_sliders.get(exp["name"])
        if var is None:
            var = tk.DoubleVar(value=0.0)
            self._sim_sliders[exp["name"]] = var
        _sim_build_expense_row(self, frame, exp, var, idx)

    _sim_refresh(self)


def _sim_open_expense_table_window(self):
    """Open the expense simulator table in a separate window."""
    existing = getattr(self, "_sim_expense_win", None)
    if existing is not None:
        try:
            if existing.winfo_exists():
                existing.deiconify()
                existing.lift()
                existing.focus_force()
                return
        except Exception:
            pass

    win = tk.Toplevel(self)
    win.title("Expense Simulator Table")
    win.configure(bg=_CARD_WHITE)
    win.resizable(True, True)
    win.grab_set()
    self._sim_expense_win = win

    p_x = self.winfo_rootx()
    p_y = self.winfo_rooty()
    p_w = self.winfo_width()
    p_h = self.winfo_height()
    w_w, w_h = 980, 620
    win.geometry(f"{w_w}x{w_h}+{p_x + (p_w - w_w)//2}+{p_y + (p_h - w_h)//2}")
    win.minsize(760, 500)

    hdr = tk.Frame(win, bg=_NAVY_DEEP)
    hdr.pack(fill="x")
    tk.Label(
        hdr,
        text="⧉  Expense Simulator Table",
        font=F(11, "bold"),
        fg=_WHITE,
        bg=_NAVY_DEEP,
        padx=16,
        pady=10,
    ).pack(side="left")
    tk.Label(
        hdr,
        text="Adjust per-expense inflation here. Client Impact in the main tab updates automatically.",
        font=F(8),
        fg="#8DA8C8",
        bg=_NAVY_DEEP,
        padx=8,
        pady=10,
    ).pack(side="left")

    # ── Search bar ─────────────────────────────────────────────────────
    search_bar = tk.Frame(win, bg=_OFF_WHITE, height=46)
    search_bar.pack(fill="x")
    search_bar.pack_propagate(False)
    tk.Label(search_bar, text="🔍", font=F(10), fg=_TXT_MUTED,
             bg=_OFF_WHITE).pack(side="left", padx=(14, 4), pady=10)
    self._sim_expense_search_var = tk.StringVar()
    _exp_search_entry = ctk.CTkEntry(
        search_bar,
        textvariable=self._sim_expense_search_var,
        width=300,
        height=28,
        corner_radius=5,
        fg_color=_WHITE,
        text_color=_TXT_NAVY,
        border_color=_BORDER_MID,
        font=FF(10),
        placeholder_text="Search expense item…",
    )
    _exp_search_entry.pack(side="left", pady=9)
    _exp_search_entry.bind(
        "<KeyRelease>",
        lambda _e: _sim_render_expense_table_rows(self),
    )
    ctk.CTkButton(
        search_bar,
        text="✕",
        width=28,
        height=28,
        corner_radius=5,
        fg_color=_OFF_WHITE,
        hover_color=_BORDER_MID,
        text_color=_TXT_MUTED,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: [
            self._sim_expense_search_var.set(""),
            _sim_render_expense_table_rows(self),
        ],
    ).pack(side="left", padx=(6, 0), pady=9)

    # ── Global inflation rate input (right side of search bar) ──────────
    tk.Frame(search_bar, bg=_BORDER_MID, width=1).pack(side="left", fill="y", padx=(16, 0), pady=8)
    tk.Label(
        search_bar,
        text="Global Rate (%):",
        font=F(9, "bold"),
        fg=_TXT_SOFT,
        bg=_OFF_WHITE,
    ).pack(side="left", padx=(14, 6), pady=9)
    _global_rate_var = tk.StringVar(value="0.0")
    _global_rate_entry = ctk.CTkEntry(
        search_bar,
        textvariable=_global_rate_var,
        width=90,
        height=28,
        corner_radius=5,
        fg_color=_WHITE,
        text_color=_TXT_NAVY,
        border_color=_BORDER_MID,
        font=FF(10),
        placeholder_text="e.g. 10.0",
    )
    _global_rate_entry.pack(side="left", pady=9)

    def _apply_global_rate_from_popup():
        try:
            pct = float(_global_rate_var.get())
        except (ValueError, TypeError):
            pct = 0.0
        if pct < 0.0:
            pct = 0.0
        _global_rate_var.set(str(pct))
        for var in self._sim_sliders.values():
            var.set(str(pct))
        _sim_render_expense_table_rows(self)

    ctk.CTkButton(
        search_bar,
        text="Apply All",
        width=82,
        height=28,
        corner_radius=5,
        fg_color=_NAVY_LIGHT,
        hover_color=_NAVY_MID,
        text_color=_WHITE,
        font=FF(9, "bold"),
        command=_apply_global_rate_from_popup,
    ).pack(side="left", padx=(6, 0), pady=9)
    ctk.CTkButton(
        search_bar,
        text="Reset",
        width=60,
        height=28,
        corner_radius=5,
        fg_color=_OFF_WHITE,
        hover_color=_BORDER_MID,
        text_color=_TXT_MUTED,
        font=FF(9, "bold"),
        border_width=1,
        border_color=_BORDER_MID,
        command=lambda: [
            _global_rate_var.set("0.0"),
            [var.set("0") for var in self._sim_sliders.values()],
            _sim_render_expense_table_rows(self),
        ],
    ).pack(side="left", padx=(4, 0), pady=9)
    _global_rate_entry.bind("<Return>", lambda _e: _apply_global_rate_from_popup())

    body = tk.Frame(win, bg=_CARD_WHITE)
    body.pack(fill="both", expand=True, padx=12, pady=12)
    sim_sb = tk.Scrollbar(body, relief="flat",
                          troughcolor=_OFF_WHITE, bg=_BORDER_LIGHT, width=8, bd=0)
    sim_sb.pack(side="right", fill="y")
    canvas = tk.Canvas(body, bg=_CARD_WHITE, highlightthickness=0,
                       yscrollcommand=sim_sb.set)
    canvas.pack(side="left", fill="both", expand=True)
    sim_sb.config(command=canvas.yview)
    frame = tk.Frame(canvas, bg=_CARD_WHITE)
    win_id = canvas.create_window((0, 0), window=frame, anchor="nw")
    frame.bind("<Configure>", lambda _e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
    canvas.bind(
        "<Enter>",
        lambda _e: canvas.bind_all(
            "<MouseWheel>",
            lambda ev: canvas.yview_scroll(int(-1 * (ev.delta / 120)), "units"),
        ),
    )
    canvas.bind("<Leave>", lambda _e: canvas.unbind_all("<MouseWheel>"))

    self._sim_canvas = canvas
    self._sim_scroll_frame = frame
    _sim_render_expense_table_rows(self)

    def _on_close():
        try:
            canvas.unbind_all("<MouseWheel>")
        except Exception:
            pass
        self._sim_canvas = None
        self._sim_scroll_frame = None
        self._sim_expense_win = None
        try:
            self._sim_expense_search_var.set("")
        except Exception:
            pass
        win.destroy()

    win.protocol("WM_DELETE_WINDOW", _on_close)


def _sim_populate(self):
    if not hasattr(self, "_sim_hdr_lbl") or not self._sim_hdr_lbl.winfo_exists():
        return

    # Cancel any in-progress build
    if getattr(self, "_sim_build_job", None):
        try:
            self._sim_hdr_lbl.after_cancel(self._sim_build_job)
        except Exception:
            pass
        self._sim_build_job = None

    filtered_data  = _lu_get_filtered_all_data(self)
    selected_inds  = getattr(self, "_sim_selected_industries", set()) or set()
    filtered_data  = _sim_filter_data_by_industry_checklist(filtered_data, selected_inds)
    q              = getattr(self, "_sim_search_var", tk.StringVar(value="")).get().strip()
    filtered_data  = _lu_filter_data_by_query(filtered_data, q)
    match_count    = len(filtered_data.get("general", []))
    match_lbl      = getattr(self, "_sim_match_lbl", None)
    filter_lbl     = getattr(self, "_sim_industry_filter_lbl", None)
    if filter_lbl is not None:
        if selected_inds:
            selected_sorted = sorted(selected_inds, key=str.lower)
            preview = ", ".join(selected_sorted[:2])
            extra_count = len(selected_sorted) - 2
            suffix = f" (+{extra_count} more)" if extra_count > 0 else ""
            filter_lbl.config(text=f"Industry filter: {preview}{suffix}")
        else:
            filter_lbl.config(text="")
    if match_lbl is not None:
        if q:
            client_names = sorted({
                (r.get("client") or "").strip()
                for r in filtered_data.get("general", [])
                if r.get("client")
            })
            if len(client_names) == 1:
                match_lbl.config(text=client_names[0][:28], bg="#4A6FA5")
            else:
                match_lbl.config(text=f"{match_count} CLIENTS MATCHED", bg="#4A6FA5")
        else:
            match_lbl.config(text="", bg=_OFF_WHITE)
    active_sectors = _lu_get_active_sectors(self)
    client         = self._lu_active_client
    is_general     = (client == GENERAL_CLIENT)
    all_clients    = filtered_data.get("clients", {})

    # If the search narrows to a single client (or exact client key match),
    # treat it as a per-client simulator view so expenses always show.
    chosen_client = None
    if q:
        ql = q.strip().lower()
        exact = next((name for name in all_clients.keys() if str(name).strip().lower() == ql), None)
        if exact:
            chosen_client = exact
        else:
            client_names = sorted({
                (r.get("client") or "").strip()
                for r in filtered_data.get("general", [])
                if r.get("client")
            })
            if len(client_names) == 1:
                chosen_client = client_names[0]

    if chosen_client and chosen_client in all_clients:
        try:
            self._lu_active_client = chosen_client
        except Exception:
            pass
        recs = [all_clients[chosen_client]]
    else:
        recs = (list(all_clients.values())
                if (is_general or active_sectors)
                else ([all_clients[client]] if client in all_clients else []))

    # ── Apply client impact search bar filter to recs ──────────────────
    # This ensures the Expense Settings popup only shows expenses that
    # belong to the clients currently visible in the Client Impact table.
    _client_search_term = str(
        getattr(self, "_sim_client_search_var", None) and
        self._sim_client_search_var.get() or ""
    ).strip().lower()
    if _client_search_term:
        recs = [
            r for r in recs
            if _client_search_term in str((r or {}).get("client") or "").lower()
        ]

    # Update header
    if q:
        self._sim_hdr_lbl.config(
            text=f"⚙️  Simulator — Search: {q[:30]}",
            fg=_LIME_MID)
    elif active_sectors:
        self._sim_hdr_lbl.config(
            text=f"⚙️  Simulator — Filtered: {' · '.join(active_sectors)}",
            fg=_LIME_MID)
    else:
        self._sim_hdr_lbl.config(text="⚙️  Inflation / Cost-Shock Simulator", fg=_WHITE)

    # Exclude synthetic summary rows ("TOTAL" / "AVERAGE") that some source
    # files store as client records — they are not real clients.
    _EXCLUDED_CLIENT_NAMES = {"total", "average"}
    recs = [
        r for r in recs
        if str((r or {}).get("client") or "").strip().lower() not in _EXCLUDED_CLIENT_NAMES
    ]

    # Use per-client net income (not total source) for simulator totals.
    net_income           = sum((r.get("net_income") or 0) for r in recs)
    self._sim_net_income = net_income
    self._sim_recs = list(recs)
    self._sim_client_page = 0

    accumulated: dict = {}
    for rec in recs:
        for exp in rec.get("expenses", []):
            name = str((exp or {}).get("name") or "").strip()
            if not name:
                continue
            try:
                total = float((exp or {}).get("total") or 0.0)
            except (TypeError, ValueError):
                total = 0.0
            if total <= 0:
                continue
            risk = str((exp or {}).get("risk") or "LOW").upper()
            if risk not in _RISK_ORDER:
                risk = "LOW"
            reason = str((exp or {}).get("reason") or "").strip()
            if name not in accumulated:
                accumulated[name] = {
                    "name": name,
                    "total": total,
                    "risk": risk,
                    "reason": reason,
                    "value_str": _fmt_value([total]),
                }
            else:
                accumulated[name]["total"] += total
                if _RISK_ORDER.get(risk, 9) < _RISK_ORDER.get(accumulated[name]["risk"], 9):
                    accumulated[name]["risk"] = risk
                    accumulated[name]["reason"] = reason
                accumulated[name]["value_str"] = _fmt_value([accumulated[name]["total"]])

    all_expenses = sorted(accumulated.values(),
                          key=lambda e: _RISK_ORDER.get(e["risk"], 9))

    # Apply row cap
    capped = False
    if len(all_expenses) > SIM_MAX_ROWS:
        all_expenses = all_expenses[:SIM_MAX_ROWS]
        capped       = True

    self._sim_expenses = all_expenses
    self._sim_expenses_capped = capped

    # Preserve previously-typed inflation rates so filter changes don't wipe them.
    _old_sliders = getattr(self, "_sim_sliders", {}) or {}
    self._sim_sliders = {}

    if not all_expenses:
        _sim_render_expense_table_rows(self)
        _sim_refresh(self)
        return

    # Pre-create DoubleVar objects, carrying forward any rate the user already typed.
    for exp in all_expenses:
        old_var = _old_sliders.get(exp["name"])
        try:
            old_val = float(old_var.get()) if old_var is not None else 0.0
        except Exception:
            old_val = 0.0
        var = tk.DoubleVar(value=old_val)
        self._sim_sliders[exp["name"]] = var

    _sim_render_expense_table_rows(self)
    _sim_refresh(self)


# ══════════════════════════════════════════════════════════════════════
#  ROW BUILDER
# ══════════════════════════════════════════════════════════════════════

def _sim_build_expense_row(self, parent, exp, var, idx):
    risk   = str(exp.get("risk") or "LOW").upper()
    if risk not in _RISK_ORDER:
        risk = "LOW"
    name = str(exp.get("name") or "Unnamed Expense")
    try:
        base_total = float(exp.get("total") or 0.0)
    except (TypeError, ValueError):
        base_total = 0.0
    row_bg = _RISK_BG.get(risk, _WHITE) if idx % 2 == 0 else _WHITE
    row    = tk.Frame(parent, bg=row_bg)
    row.pack(fill="x")
    for ci, (_title, min_px, _wt) in enumerate(SIM_TABLE_COLUMNS):
        row.grid_columnconfigure(ci, weight=1, minsize=min_px, uniform="sim_col")

    tk.Label(row, text=name, font=F(9, "bold"),
             fg=_TXT_NAVY, bg=row_bg, anchor="w", padx=8, pady=6
             ).grid(row=0, column=0, sticky="ew")
    tk.Label(row, text=risk, font=F(7, "bold"),
             fg=_RISK_COLOR.get(risk, _TXT_SOFT),
             bg=_RISK_BADGE_BG.get(risk, _OFF_WHITE),
             anchor="center", justify="center",
             padx=10, pady=3).grid(row=0, column=1, padx=4, pady=6, sticky="")
    tk.Label(row, text=f"₱{base_total:,.2f}" if base_total > 0 else "—",
             font=F(9), fg=_TXT_NAVY, bg=row_bg, anchor="center", justify="center", padx=6
             ).grid(row=0, column=2, sticky="ew")

    rate_entry = ctk.CTkEntry(row, textvariable=var, width=80, height=26, corner_radius=4,
                              font=FF(9), fg_color=_WHITE, text_color=_TXT_NAVY,
                              border_color=_RISK_COLOR.get(risk, _BORDER_MID),
                              placeholder_text="0")
    rate_entry.grid(row=0, column=3, padx=8, pady=6)
    rate_entry.bind("<Return>",   lambda e, ex=exp: _sim_on_slide(self, ex, var.get()))
    rate_entry.bind("<FocusOut>", lambda e, ex=exp: _sim_on_slide(self, ex, var.get()))

    extra_lbl = tk.Label(row, text="—", font=F(9), fg=_ACCENT_RED,
                         bg=row_bg, anchor="center", justify="center", padx=6)
    extra_lbl.grid(row=0, column=4, sticky="ew")
    sim_lbl = tk.Label(row, text="—", font=F(9, "bold"), fg=_TXT_NAVY,
                       bg=row_bg, anchor="center", justify="center", padx=6)
    sim_lbl.grid(row=0, column=5, sticky="ew")

    var._extra_lbl = extra_lbl
    var._sim_lbl   = sim_lbl
    var._base      = base_total
    tk.Frame(parent, bg=_BORDER_LIGHT, height=1).pack(fill="x")


# ══════════════════════════════════════════════════════════════════════
#  INTERACTION CALLBACKS
# ══════════════════════════════════════════════════════════════════════

def _sim_on_slide(self, exp, value):
    try:
        pct = float(value)
    except (ValueError, TypeError):
        pct = 0.0
    if pct < 0.0:
        pct = 0.0
    self._sim_sliders[exp["name"]].set(str(pct))
    _sim_refresh(self)


def _sim_apply_global(self):
    try:
        pct = float(self._sim_global_var.get())
    except (ValueError, TypeError):
        pct = 0.0
    if pct < 0.0:
        pct = 0.0
    if not self._sim_sliders and self._lu_all_data:
        _sim_populate(self)
    for var in self._sim_sliders.values():
        var.set(str(pct))
    _sim_refresh(self)


def _sim_reset(self):
    self._sim_global_var.set("0")
    if not self._sim_sliders and self._lu_all_data:
        _sim_populate(self)
    for var in self._sim_sliders.values():
        var.set("0")
    _sim_refresh(self)


def _sim_refresh(self):
    base_total = sim_total = 0.0
    for exp in getattr(self, "_sim_expenses", []):
        pct = 0.0
        var = self._sim_sliders.get(exp["name"])
        if var:
            try:
                pct = float(var.get())
            except (ValueError, TypeError):
                pass
        base  = exp["total"]
        extra = base * pct / 100.0
        sim   = base + extra
        base_total += base
        sim_total  += sim
        if var and hasattr(var, "_extra_lbl"):
            try:
                var._extra_lbl.config(
                    text=f"+₱{extra:,.2f}" if extra > 0 else "—",
                    fg=_ACCENT_RED if extra > 0 else _TXT_MUTED)
                var._sim_lbl.config(
                    text=f"₱{sim:,.2f}" if base > 0 else "—", fg=_TXT_NAVY)
            except Exception:
                pass

    increase   = sim_total - base_total
    net_income = getattr(self, "_sim_net_income", 0.0) or 0.0
    surplus    = net_income - sim_total

    if hasattr(self, "_sim_lbl_base"):
        try:
            self._sim_lbl_income.config(
                text=f"₱{net_income:,.2f}" if net_income else "—",
                fg=_ACCENT_SUCCESS)
            self._sim_lbl_base.config(
                text=f"₱{base_total:,.2f}" if base_total else "—")
            self._sim_lbl_sim.config(
                text=f"₱{sim_total:,.2f}" if base_total else "—")
            self._sim_lbl_inc.config(
                text=f"+₱{increase:,.2f}" if increase > 0 else "₱0.00",
                fg=_ACCENT_RED if increase > 0 else _TXT_NAVY)
            if net_income:
                surplus_txt = f"{'▲' if surplus >= 0 else '▼'} ₱{abs(surplus):,.2f}"
                self._sim_lbl_surplus.config(
                    text=surplus_txt,
                    fg=_ACCENT_SUCCESS if surplus >= 0 else _ACCENT_RED)
            else:
                self._sim_lbl_surplus.config(text="—", fg=_TXT_MUTED)
        except Exception:
            pass

    if hasattr(self, "_sim_income_lbl"):
        try:
            if net_income:
                self._sim_income_lbl.config(
                    text=f"TOTAL NET INCOME  ₱{net_income:,.2f}",
                    fg=_LIME_MID)
                self._sim_surplus_lbl.config(
                    text=(f"SURPLUS  ₱{surplus:,.2f}" if surplus >= 0
                          else f"DEFICIT  ▲ ₱{abs(surplus):,.2f}"),
                    fg=_LIME_MID if surplus >= 0 else _ACCENT_RED)
            else:
                self._sim_income_lbl.config(
                    text="TOTAL NET INCOME  —  Load a file to begin",
                    fg=_TXT_MUTED)
                self._sim_surplus_lbl.config(text="", fg=_LIME_MID)
        except Exception:
            pass

    _sim_draw_chart(self)
    _sim_refresh_client_table(self)


_NEGATIVE_INCOME_LABEL = "HIGH (NEGATIVE INCOME)"
_NEGATIVE_INCOME_PCT_DISPLAY = "N/A (Negative Income)"


def _sim_pct_net_to_amort_label(pct: float, ranges: dict | None = None) -> str:
    """
    Risk Label rules for the simulator client table.
    Ranges default to _SIM_DEFAULT_RISK_RANGES but can be overridden at runtime
    via self._sim_risk_ranges (passed in as the `ranges` argument).

    Each range entry: {"LOW": (min, max), "MEDIUM": (min, max), "HIGH": (min, max)}
    Boundaries are inclusive on both ends. HIGH max may be float('inf').

    Special sentinels:
      pct == -2.0  → zero amortization; returns "N/A" (not meaningful to assess risk).
      pct <  0     → negative net income; returns _NEGATIVE_INCOME_LABEL.
    Falls back to LOW if pct matches no range.
    """
    if ranges is None:
        ranges = _SIM_DEFAULT_RISK_RANGES
    try:
        p = float(pct)
    except Exception:
        return "LOW"
    if p == -2.0:
        return "N/A"
    if p < 0:
        return _NEGATIVE_INCOME_LABEL
    for label in ("HIGH", "MEDIUM", "LOW"):
        lo, hi = ranges.get(label, (0.0, -1.0))
        if lo <= p <= hi:
            return label
    return "LOW"


def _sim_build_risk_reasoning(risk_label: str, pct_net_to_am: float) -> str:
    """Build user-facing simulator risk explanation text."""
    label = str(risk_label or "LOW").upper()
    if label == "N/A":
        return (
            "The client has no recorded amortization (₱0.00). "
            "Risk cannot be meaningfully assessed — marked N/A."
        )
    if "NEGATIVE" in label:
        return (
            "The client has NEGATIVE simulated net income. "
            "Amortization % cannot be meaningfully computed. "
            "Immediate review required."
        )
    base = (
        f"The client is {label} risk because they have "
        f"{float(pct_net_to_am):.1f}% of Net Income to Amortization."
    )
    if label == "HIGH":
        return base + " Please review carefully — this client might have a special loan case."
    return base


def _sim_amount_for_expense(self, base_amount: float, expense_name: str) -> tuple[float, float]:
    """Return (extra_cost, simulated_amount) for an expense name using current slider %."""
    pct = 0.0
    var = getattr(self, "_sim_sliders", {}).get(expense_name)
    if var is not None:
        try:
            pct = float(var.get() or 0.0)
        except Exception:
            pct = 0.0
    if pct < 0.0:
        pct = 0.0
    base = float(base_amount or 0.0)
    extra = base * pct / 100.0
    return (extra, base + extra)


def _sim_refresh_client_table(self):
    tree = getattr(self, "_sim_client_tree", None)
    if tree is None:
        return
    try:
        if not tree.winfo_exists():
            return
    except Exception:
        return

    recs_all = list(getattr(self, "_sim_recs", []) or [])
    page = int(getattr(self, "_sim_client_page", 0) or 0)

    # Used by click handler to reliably map a clicked Treeview row
    # back to the full client name (even if the visible label is truncated).
    self._sim_iid_to_client = {}
    self._sim_client_metrics_by_name = {}

    # Apply client name search filter
    _search_term = str(getattr(self, "_sim_client_search_var", None) and
                       self._sim_client_search_var.get() or "").strip().lower()
    if _search_term:
        recs_all = [
            r for r in recs_all
            if _search_term in str((r or {}).get("client") or "").lower()
        ]
    for iid in tree.get_children(""):
        tree.delete(iid)

    # Reset hover tracking on every repopulate so stale iids don't persist.
    try:
        tree._sim_hovered_iid_ref[0] = None
    except AttributeError:
        pass
    if not recs_all:
        tree.insert("", "end", values=("—",) + ("—",) * (len(SIM_CLIENT_TABLE_COLUMNS) - 1), tags=("NA",))
        _sim_update_client_pagination_ui(self, total_rows=0)
        return

    def _money(v: float | None) -> str:
        try:
            if v is None:
                return "—"
            return f"₱{float(v):,.2f}"
        except Exception:
            return "—"

    def _pct(v: float | None) -> str:
        try:
            if v is None:
                return "—"
            fv = float(v)
            if fv < 0:
                return _NEGATIVE_INCOME_PCT_DISPLAY
            return f"{fv:.1f}%"
        except Exception:
            return "—"

    rows = []
    for rec in recs_all:
        name = str((rec or {}).get("client") or "").strip() or "—"
        industry = str((rec or {}).get("industry") or "").strip() or "—"
        base_net = float((rec or {}).get("net_income") or 0.0)
        current_am = float((rec or {}).get("current_amort") or 0.0)

        # Build per-client base expense map from parsed expenses (same names used in simulator sliders).
        base_by_name: dict[str, float] = {}
        base_total_exp = 0.0
        for exp in (rec or {}).get("expenses", []) or []:
            nm = str((exp or {}).get("name") or "").strip()
            if not nm:
                continue
            try:
                amt = float((exp or {}).get("total") or 0.0)
            except Exception:
                amt = 0.0
            if amt <= 0:
                continue
            base_by_name[nm] = base_by_name.get(nm, 0.0) + amt
            base_total_exp += amt

        # Apply current slider ramp-ups to THIS client's expenses.
        extra_total = 0.0
        for nm, base_amt in base_by_name.items():
            extra, _sim = _sim_amount_for_expense(self, base_amt, nm)
            extra_total += extra

        sim_total_exp = base_total_exp + extra_total

        pct_inc = (extra_total / base_total_exp * 100.0) if base_total_exp > 0 else (0.0 if extra_total <= 0 else 100.0)

        # Simulated net income = Base net income minus the simulated cost increase
        sim_net_income = base_net - extra_total

        # % Net → Amort uses simulated net income (Total Current Amort / Total Net Income Simulated)
        if current_am == 0:
            pct_net_to_am = -2.0  # sentinel: no amortization — risk is N/A
        elif sim_net_income <= 0:
            pct_net_to_am = -1.0  # sentinel: negative income
        else:
            pct_net_to_am = (current_am / sim_net_income) * 100.0

        sim_risk = _sim_pct_net_to_amort_label(pct_net_to_am, getattr(self, "_sim_risk_ranges", None))
        risk_reasoning = _sim_build_risk_reasoning(sim_risk, pct_net_to_am)
        r = {
            "client": name,
            "industry": industry,
            "base_total_expenses": base_total_exp,
            "sim_total_expenses": sim_total_exp,
            "net_income": base_net,
            "sim_net_income": sim_net_income,
            "pct_increase": pct_inc,
            "sim_increase": extra_total,
            "current_amort": current_am,
            "pct_net_to_amort": pct_net_to_am,
            "sim_risk_label": sim_risk,
            "risk_reasoning": risk_reasoning,
        }
        rows.append(r)
        self._sim_client_metrics_by_name[name] = r

    # Sort by highest % Net → Amort first (descending).
    rows.sort(key=lambda r: -(r["pct_net_to_amort"] or 0.0))

    total = len(rows)
    max_page = max(0, (total - 1) // SIM_CLIENT_PAGE_SIZE) if total else 0
    if page < 0:
        page = 0
    if page > max_page:
        page = max_page
    self._sim_client_page = page
    start = page * SIM_CLIENT_PAGE_SIZE
    end = min(start + SIM_CLIENT_PAGE_SIZE, total)
    page_rows = rows[start:end]
    _sim_update_client_pagination_ui(self, total_rows=total)

    for idx_in_page, r in enumerate(page_rows):
        # Keep full client name for matching with Analysis tab.
        full_client_name = r.get("client") or "—"
        iid = f"sim_client_{start + idx_in_page}"
        vals = (
            r["client"][:60],
            r["industry"],
            _money(r["base_total_expenses"]),
            _money(r["sim_total_expenses"]),
            _money(r["net_income"]),
            _money(r["sim_net_income"]),
            _pct(r["pct_increase"]),
            _money(r["sim_increase"]),
            _money(r["current_amort"]),
            _pct(r["pct_net_to_amort"]),
            r["sim_risk_label"],
            r["risk_reasoning"],
        )
        _lbl = r["sim_risk_label"]
        if _lbl == "HIGH":
            tag = "HIGH"
        elif "NEGATIVE" in str(_lbl).upper():
            tag = "HIGH_NEG"
        elif _lbl in ("MEDIUM", "LOW"):
            tag = _lbl
        else:
            tag = "NA"
        self._sim_iid_to_client[iid] = str(full_client_name)
        tree.insert("", "end", iid=iid, values=vals, tags=(tag,))


def _sim_show_table_view(self):
    """
    Hide the single-client detail view and restore the full treeview table.
    Called by the Back button and whenever the data is refreshed.
    """
    detail = getattr(self, "_sim_client_detail_view", None)
    table  = getattr(self, "_sim_client_table_view",  None)

    if detail is not None:
        try:
            detail.pack_forget()
        except Exception:
            pass
    if table is not None:
        try:
            table.pack(fill="x", expand=False)
        except Exception:
            pass


def _sim_show_detail_view(self, client_name: str):
    """
    Hide the treeview table, show the detail panel for *client_name*.
    Called when a row is clicked.
    """
    table  = getattr(self, "_sim_client_table_view",  None)
    detail = getattr(self, "_sim_client_detail_view", None)

    if table is not None:
        try:
            table.pack_forget()
        except Exception:
            pass
    if detail is not None:
        try:
            detail.pack(fill="x", expand=False, padx=0, pady=0)
        except Exception:
            pass

    # Render the detail content (Back button is rendered inside by _sim_show_client_details)
    _sim_show_client_details(self, client_name)


def _sim_on_client_impact_row_activated(self, event=None):
    """
    Mirror Analysis treeview behavior:
    when a user clicks a Risk Simulator client row, open that client's
    full details in the Analysis tab.
    """
    tree = getattr(self, "_sim_client_tree", None)
    if tree is None:
        return

    iid = None
    try:
        if event is not None and hasattr(event, "y"):
            iid = tree.identify_row(event.y)
        else:
            sel = tree.selection()
            iid = sel[0] if sel else None
    except Exception:
        iid = None

    client_name = getattr(self, "_sim_iid_to_client", {}).get(iid)
    if not client_name or client_name == "—":
        return

    # Keep the Analysis dropdown synchronized (so clicking Analysis tab shows
    # the correct selected client), but do the "all columns" display inside
    # the Simulator tab as requested.
    try:
        if getattr(self, "_lu_client_var", None) is not None:
            self._lu_client_var.set(client_name)
    except Exception:
        pass

    try:
        if hasattr(self, "_lu_on_client_change"):
            self._lu_on_client_change(client_name)
    except Exception:
        pass

    _sim_show_detail_view(self, client_name)


def _sim_show_client_details(self, client_name: str):
    """
    Render the full single-client panel inside _sim_client_detail_view.
    Called by _sim_show_detail_view after the table is hidden.
    """
    detail = getattr(self, "_sim_client_detail_frame", None)
    if detail is None:
        return

    # Wipe previous content
    for w in detail.winfo_children():
        try:
            w.destroy()
        except Exception:
            pass

    if not client_name or client_name == "—":
        return

    metrics = getattr(self, "_sim_client_metrics_by_name", {}) or {}
    m = metrics.get(client_name) or {}

    # Prefer full record from LU core output
    rec = (getattr(self, "_lu_all_data", None) or {}).get("clients", {}).get(client_name)
    if not rec:
        for rr in getattr(self, "_sim_recs", []) or []:
            if str((rr or {}).get("client") or "").strip() == client_name:
                rec = rr
                break

    sim_label = str(m.get("sim_risk_label") or "LOW").upper()
    badge_bg  = _RISK_BADGE_BG.get(sim_label, _OFF_WHITE)
    badge_fg  = _RISK_COLOR.get(sim_label, _TXT_MUTED)

    # ── Top-accent color per risk (matches the card screenshot) ────────
    _ACCENT_BAR = {
        "HIGH":   _ACCENT_RED,      # red
        "MEDIUM": _ACCENT_GOLD,     # gold/amber
        "LOW":    _ACCENT_SUCCESS,  # green
    }
    accent_color = _ACCENT_BAR.get(sim_label, _LIME_MID)

    # ── Single unified card (one accent bar, hero + metrics together) ──
    client_card = tk.Frame(detail, bg=_NAVY_DEEP)
    client_card.pack(fill="x")

    # Single colored top accent bar only
    tk.Frame(client_card, bg=accent_color, height=4).pack(fill="x")

    hero_inner = tk.Frame(client_card, bg=_NAVY_DEEP)
    hero_inner.pack(fill="x", padx=24, pady=16)

    # Back to Table button — lives in the same container as the client name
    ctk.CTkButton(
        hero_inner,
        text="◄  Back to Table",
        width=130,
        height=28,
        corner_radius=6,
        fg_color=_NAVY_LIGHT,
        hover_color=_NAVY_MID,
        text_color=_WHITE,
        font=FF(9, "bold"),
        command=lambda: _sim_show_table_view(self),
    ).pack(side="right", padx=(0, 4), pady=(0, 8))

    tk.Label(
        hero_inner,
        text=client_name,
        font=F(13, "bold"),
        fg=_WHITE,
        bg=_NAVY_DEEP,
        anchor="w",
        justify="left",
        wraplength=900,
    ).pack(anchor="w")

    badge_row = tk.Frame(hero_inner, bg=_NAVY_DEEP)
    badge_row.pack(anchor="w", pady=(8, 0))
    badge = tk.Frame(badge_row, bg=badge_bg,
                     highlightbackground=badge_fg, highlightthickness=1)
    badge.pack(side="left")
    tk.Label(
        badge,
        text=f"  {sim_label} RISK  ",
        font=F(9, "bold"),
        fg=badge_fg,
        bg=badge_bg,
        padx=10,
        pady=4,
    ).pack()

    # Reasoning beside badge
    tk.Label(
        badge_row,
        text=str(m.get("risk_reasoning") or ""),
        font=F(8),
        fg="#8DAACC",
        bg=_NAVY_DEEP,
        anchor="w",
        justify="left",
        wraplength=800,
        padx=14,
    ).pack(side="left", fill="x", expand=True)

    # Subtle inner divider — no second accent bar
    tk.Frame(client_card, bg="#1E3A5F", height=1).pack(fill="x", padx=0)

    summary = tk.Frame(client_card, bg=_NAVY_DEEP)
    summary.pack(fill="x")


    def _money(v):
        try:
            return f"₱{float(v):,.2f}"
        except Exception:
            return "—"

    def _pct(v):
        try:
            fv = float(v)
            if fv < 0:
                return _NEGATIVE_INCOME_PCT_DISPLAY
            return f"{fv:.1f}%"
        except Exception:
            return "—"

    metrics_pairs = [
        ("Total Expenses (Base)", _money(m.get("base_total_expenses"))),
        ("Total Expenses (Sim)",  _money(m.get("sim_total_expenses"))),
        ("Net Income (Base)",     _money(m.get("net_income"))),
        ("Net Income (Sim)",      _money(m.get("sim_net_income"))),
        ("% Expense Increase",    _pct(m.get("pct_increase"))),
        ("Simulated Increase",    _money(m.get("sim_increase"))),
        ("Current Amort",         _money(m.get("current_amort"))),
        ("% Net → Amort (Sim)",   _pct(m.get("pct_net_to_amort"))),
    ]
    for lbl, val in metrics_pairs:
        c = tk.Frame(summary, bg=_NAVY_DEEP)
        c.pack(side="left", padx=14, pady=12)
        tk.Label(c, text=lbl,  font=F(7),          fg="#8DAACC",  bg=_NAVY_DEEP).pack(anchor="w")
        tk.Label(c, text=val,  font=F(10, "bold"),  fg=_WHITE,     bg=_NAVY_DEEP).pack(anchor="w")

    tk.Frame(detail, bg=_BORDER_MID, height=1).pack(fill="x")

    # ── Full LU record — section title ───────────────────────────────
    rec_hdr = tk.Frame(detail, bg=_OFF_WHITE)
    rec_hdr.pack(fill="x")
    tk.Label(
        rec_hdr,
        text="Full Client Record  —  All Columns",
        font=F(9, "bold"),
        fg=_TXT_SOFT,
        bg=_OFF_WHITE,
        anchor="w",
        padx=20,
        pady=8,
    ).pack(anchor="w")
    tk.Frame(detail, bg=_BORDER_LIGHT, height=1).pack(fill="x")

    if not rec:
        tk.Label(
            detail,
            text="No full LU record found for this client.",
            font=F(9),
            fg=_ACCENT_RED,
            bg=_CARD_WHITE,
            wraplength=1100,
            justify="left",
            padx=20,
            pady=20,
        ).pack(anchor="w")
        return

    row_idx = 0
    for _cid, heading, field, _w, _a, kind in LU_CLIENT_TREE_SPEC:
        if field not in rec:
            continue
        raw_val = rec.get(field)
        if raw_val is None or str(raw_val).strip() == "":
            continue
        value = lu_format_lu_cell(rec, field, kind, text_limit=500)

        bg  = _WHITE if row_idx % 2 == 0 else _OFF_WHITE
        row = tk.Frame(detail, bg=bg)
        row.pack(fill="x", padx=0, pady=0)

        tk.Label(
            row,
            text=heading,
            font=F(8, "bold"),
            fg=_TXT_SOFT,
            bg=bg,
            width=28,
            anchor="nw",
            justify="left",
            padx=20,
            pady=7,
        ).pack(side="left", fill="y")

        tk.Label(
            row,
            text=value,
            font=F(9),
            fg=_TXT_NAVY,
            bg=bg,
            anchor="w",
            justify="left",
            wraplength=1000,
            padx=8,
            pady=7,
        ).pack(side="left", fill="x", expand=True)

        tk.Frame(detail, bg=_BORDER_LIGHT, height=1).pack(fill="x", pady=0)
        row_idx += 1

    # Bottom spacer
    tk.Frame(detail, bg=_CARD_WHITE, height=24).pack(fill="x")


def _sim_write_export_settings_sheet(ws, *, self_app, fname, generated_at, exported_by, export_scope_note):
    """
    Write a rich "Export settings" sheet that mirrors the on-screen layout in the
    screenshot — plus two new sections specific to the Risk Simulator:

      • Expense modifications  — every expense item whose slider ≠ 0, shown as
                                  "Fuel and Transportation — +33% increase"
      • Industry checklist     — which industries were active in the simulator filter
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return   # silent — caller will catch missing openpyxl separately

    # ── Palette (matches the navy theme in the screenshot) ────────────
    _NAV  = "1A3A5C"   # deep navy section banners
    _WHT  = "FFFFFF"
    _HDR_BG = "EAF2FB"  # light blue sub-header
    _HDR_FG = "1A3A5C"
    _ROW_ODD  = "FFFFFF"
    _ROW_EVEN = "F4F8FD"
    _LBL_FG   = "1A3A5C"   # bold label text
    _VAL_FG   = "374151"   # regular value text
    _NONE_FG  = "9CA3AF"   # grey for "— None —" placeholders
    _MOD_POS  = "1E8449"   # green for positive % changes
    _MOD_NEG  = "C0392B"   # red  for negative % changes
    _BORDER   = "D5DCE4"

    _thin  = Side(border_style="thin",   color=_BORDER)
    _med   = Side(border_style="medium", color=_NAV)
    _grid  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _bot_m = Border(left=_thin, right=_thin, top=_thin, bottom=_med)

    COL_A = 1   # label column
    COL_B = 2   # value column
    W_A   = 30  # label column width
    W_B   = 90  # value column width

    ws.column_dimensions[get_column_letter(COL_A)].width = W_A
    ws.column_dimensions[get_column_letter(COL_B)].width = W_B

    _row = [1]   # mutable row pointer

    def _next():
        r = _row[0]; _row[0] += 1; return r

    def _banner(text, color=_NAV):
        r = _next()
        ws.merge_cells(start_row=r, start_column=COL_A, end_row=r, end_column=COL_B)
        c = ws.cell(r, COL_A, text)
        c.fill      = PatternFill("solid", fgColor=color)
        c.font      = Font(bold=True, size=11, color=_WHT, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = _bot_m
        ws.row_dimensions[r].height = 22
        return r

    def _sub_header(label_a, label_b=""):
        r = _next()
        for ci, txt in ((COL_A, label_a), (COL_B, label_b)):
            c = ws.cell(r, ci, txt)
            c.fill      = PatternFill("solid", fgColor=_HDR_BG)
            c.font      = Font(bold=True, size=9, color=_HDR_FG, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border    = _grid
        ws.row_dimensions[r].height = 16
        return r

    def _kv(label, value, label_bold=True, val_fg=_VAL_FG, row_idx=None, wrap=True):
        r = row_idx if row_idx is not None else _next()
        lc = ws.cell(r, COL_A, label)
        lc.font      = Font(bold=label_bold, size=9, color=_LBL_FG, name="Calibri")
        lc.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
        lc.fill      = PatternFill("solid", fgColor=_ROW_ODD)
        lc.border    = _grid

        vc = ws.cell(r, COL_B, value)
        vc.font      = Font(size=9, color=val_fg, name="Calibri")
        vc.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=wrap)
        vc.fill      = PatternFill("solid", fgColor=_ROW_ODD)
        vc.border    = _grid
        ws.row_dimensions[r].height = 16
        return r

    def _spacer():
        r = _next()
        for ci in (COL_A, COL_B):
            c = ws.cell(r, ci, "")
            c.fill   = PatternFill("solid", fgColor=_ROW_ODD)
            c.border = _grid
        ws.row_dimensions[r].height = 8
        return r

    # ══════════════════════════════════════════════════════════════════
    #  SECTION 1 — Settings for this export
    # ══════════════════════════════════════════════════════════════════
    _banner("Settings for this export")
    _kv("When exported",       generated_at)
    _kv("Exported by",         exported_by)
    _kv("Portfolio file used", fname)
    _kv("What this workbook is", export_scope_note)

    # Risk rules in effect
    _risk_ranges = getattr(self_app, "_sim_risk_ranges", None) or {}
    _low_r  = _risk_ranges.get("LOW",    (1.0,  35.0))
    _med_r  = _risk_ranges.get("MEDIUM", (36.0, 70.0))
    _high_r = _risk_ranges.get("HIGH",   (70.01, float("inf")))
    def _fmt_range(lo, hi):
        return f"{lo:.0f}% – {hi:.0f}%" if hi != float("inf") else f"≥ {lo:.0f}%"
    risk_rules_text = (
        f"LOW:     {_fmt_range(*_low_r)}\n"
        f"MEDIUM:  {_fmt_range(*_med_r)}\n"
        f"HIGH:    {_fmt_range(*_high_r)}"
    )
    r_risk = _next()
    lc = ws.cell(r_risk, COL_A, "Risk thresholds (% Net → Amort)")
    lc.font      = Font(bold=True, size=9, color=_LBL_FG, name="Calibri")
    lc.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    lc.fill      = PatternFill("solid", fgColor=_ROW_ODD); lc.border = _grid
    vc = ws.cell(r_risk, COL_B, risk_rules_text)
    vc.font      = Font(size=9, color=_VAL_FG, name="Calibri")
    vc.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=True)
    vc.fill      = PatternFill("solid", fgColor=_ROW_ODD); vc.border = _grid
    ws.row_dimensions[r_risk].height = 48

    _spacer()

    # ── IMPORTANT NOTICE — Amortization assumption disclaimer ─────────
    r_notice = _next()
    ws.merge_cells(start_row=r_notice, start_column=COL_A, end_row=r_notice, end_column=COL_B)
    _nc2 = ws.cell(r_notice, COL_A,
        "⚠  IMPORTANT NOTICE: Not all Net Income and Amortization figures shown in this report "
        "are guaranteed to be accurate. The simulator ASSUMES that ALL clients have an "
        "amortization obligation. This may not reflect reality — some clients may have NO active "
        "amortization. Please DOUBLE-CHECK each client record to verify whether the client "
        "actually has amortization before acting on any risk label or percentage shown here."
    )
    _nc2.fill      = PatternFill("solid", fgColor="FFF3CD")
    _nc2.font      = Font(bold=True, size=9, color="856404", name="Calibri")
    _nc2.alignment = Alignment(horizontal="left", vertical="top", indent=2, wrap_text=True)
    _nc2.border    = Border(
        left=Side(border_style="medium", color="856404"),
        right=Side(border_style="medium", color="856404"),
        top=Side(border_style="medium", color="856404"),
        bottom=Side(border_style="medium", color="856404"),
    )
    ws.row_dimensions[r_notice].height = 60

    _spacer()

    # ══════════════════════════════════════════════════════════════════
    #  SECTION 2 — Expense modifications
    # ══════════════════════════════════════════════════════════════════
    _banner("Expense modifications applied in this simulation")

    # Collect all expense names + their slider values
    sliders   = getattr(self_app, "_sim_sliders",  {}) or {}
    expenses  = getattr(self_app, "_sim_expenses", []) or []
    # Build ordered list of (name, pct) for every expense with a non-zero slider
    modified = []
    zero_list = []
    for exp in expenses:
        name = str((exp or {}).get("name") or "").strip()
        if not name:
            continue
        var = sliders.get(name)
        pct = 0.0
        if var:
            try:
                pct = float(var.get() or 0)
            except Exception:
                pct = 0.0
        if pct != 0.0:
            modified.append((name, pct))
        else:
            zero_list.append(name)

    # Also pick up any slider keys not in _sim_expenses list
    for nm, var in sliders.items():
        if nm not in {e.get("name","") for e in expenses}:
            try:
                pct = float(var.get() or 0)
            except Exception:
                pct = 0.0
            if pct != 0.0:
                modified.append((nm, pct))

    _sub_header("Expense Item", "Change Applied")

    if modified:
        modified.sort(key=lambda x: -abs(x[1]))  # largest change first
        for idx, (nm, pct) in enumerate(modified):
            r = _next()
            bg = _ROW_ODD if idx % 2 == 0 else _ROW_EVEN
            change_str = f"+{pct:.1f}% increase" if pct > 0 else f"{pct:.1f}% decrease"
            val_color  = _MOD_POS if pct > 0 else _MOD_NEG

            lc = ws.cell(r, COL_A, nm)
            lc.fill      = PatternFill("solid", fgColor=bg)
            lc.font      = Font(size=9, color=_LBL_FG, name="Calibri")
            lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            lc.border    = _grid

            vc = ws.cell(r, COL_B, change_str)
            vc.fill      = PatternFill("solid", fgColor=bg)
            vc.font      = Font(bold=True, size=9, color=val_color, name="Calibri")
            vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            vc.border    = _grid
            ws.row_dimensions[r].height = 16
    else:
        r = _next()
        ws.merge_cells(start_row=r, start_column=COL_A, end_row=r, end_column=COL_B)
        c = ws.cell(r, COL_A, "— No expense modifications — all sliders at 0% —")
        c.fill      = PatternFill("solid", fgColor=_ROW_ODD)
        c.font      = Font(italic=True, size=9, color=_NONE_FG, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c.border    = _grid
        ws.row_dimensions[r].height = 16

    # Summary count row
    _total_exp = len(modified) + len(zero_list)
    r_sum = _next()
    ws.merge_cells(start_row=r_sum, start_column=COL_A, end_row=r_sum, end_column=COL_B)
    c_sum = ws.cell(r_sum, COL_A,
        f"  {len(modified)} of {_total_exp} expense item(s) modified   |   "
        f"{len(zero_list)} at 0% (no change)"
    )
    c_sum.fill      = PatternFill("solid", fgColor=_HDR_BG)
    c_sum.font      = Font(italic=True, size=8, color=_HDR_FG, name="Calibri")
    c_sum.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c_sum.border    = _grid
    ws.row_dimensions[r_sum].height = 14

    _spacer()

    # Final spacer
    _spacer()


def _sim_write_settings_sheet(ws_settings, *, self_app, row_offset: int = 0):
    """
    Write a dedicated 'Settings' section that documents:
      1. How the % Net Income to Amortization is calculated
      2. How each Risk label (LOW / MEDIUM / HIGH) is achieved
      3. Expense modifications applied in the simulation
      4. Industries used in the Industry Checklist

    row_offset: if > 0, content is appended starting at that row (used when
    this function writes onto an existing sheet after export-settings content).
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return

    # ── Palette (matches the navy theme) ────────────────────────────────
    _NAV      = "1A3A5C"
    _WHT      = "FFFFFF"
    _HDR_BG   = "EAF2FB"
    _HDR_FG   = "1A3A5C"
    _ROW_ODD  = "FFFFFF"
    _ROW_EVEN = "F4F8FD"
    _LBL_FG   = "1A3A5C"
    _VAL_FG   = "374151"
    _NONE_FG  = "9CA3AF"
    _MOD_POS  = "1E8449"
    _MOD_NEG  = "C0392B"
    _GOLD     = "9A6700"
    _BORDER   = "D5DCE4"
    _EXPL_BG  = "F0F7FF"   # light blue tint for explanation rows
    _EXPL_FG  = "2C5282"   # medium navy for explanation text

    _thin  = Side(border_style="thin",   color=_BORDER)
    _med   = Side(border_style="medium", color=_NAV)
    _grid  = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _bot_m = Border(left=_thin, right=_thin, top=_thin, bottom=_med)

    COL_A = 1
    COL_B = 2
    COL_C = 3
    ws_settings.column_dimensions[get_column_letter(COL_A)].width = 32
    ws_settings.column_dimensions[get_column_letter(COL_B)].width = 40
    ws_settings.column_dimensions[get_column_letter(COL_C)].width = 60

    _row = [max(1, int(row_offset) + 1)]

    def _next():
        r = _row[0]; _row[0] += 1; return r

    def _banner(text, color=_NAV):
        r = _next()
        ws_settings.merge_cells(start_row=r, start_column=COL_A, end_row=r, end_column=COL_C)
        c = ws_settings.cell(r, COL_A, text)
        c.fill      = PatternFill("solid", fgColor=color)
        c.font      = Font(bold=True, size=11, color=_WHT, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = _bot_m
        ws_settings.row_dimensions[r].height = 22
        return r

    def _sub_header(col_a_txt, col_b_txt="", col_c_txt=""):
        r = _next()
        for ci, txt in ((COL_A, col_a_txt), (COL_B, col_b_txt), (COL_C, col_c_txt)):
            c = ws_settings.cell(r, ci, txt)
            c.fill      = PatternFill("solid", fgColor=_HDR_BG)
            c.font      = Font(bold=True, size=9, color=_HDR_FG, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            c.border    = _grid
        ws_settings.row_dimensions[r].height = 16
        return r

    def _row3(col_a_val, col_b_val, col_c_val, bg=_ROW_ODD,
              a_bold=False, b_bold=False, c_bold=False,
              a_fg=_LBL_FG, b_fg=_VAL_FG, c_fg=_VAL_FG,
              wrap_c=True, height=16):
        r = _next()
        for ci, txt, bold, fg, wrap in (
            (COL_A, col_a_val, a_bold, a_fg, False),
            (COL_B, col_b_val, b_bold, b_fg, False),
            (COL_C, col_c_val, c_bold, c_fg, wrap_c),
        ):
            c = ws_settings.cell(r, ci, txt)
            c.fill      = PatternFill("solid", fgColor=bg)
            c.font      = Font(bold=bold, size=9, color=fg, name="Calibri")
            c.alignment = Alignment(horizontal="left", vertical="top", indent=1, wrap_text=wrap)
            c.border    = _grid
        ws_settings.row_dimensions[r].height = height
        return r

    def _expl_row(text, height=30):
        """Full-width explanation row with light blue background."""
        r = _next()
        ws_settings.merge_cells(start_row=r, start_column=COL_A, end_row=r, end_column=COL_C)
        c = ws_settings.cell(r, COL_A, text)
        c.fill      = PatternFill("solid", fgColor=_EXPL_BG)
        c.font      = Font(italic=True, size=9, color=_EXPL_FG, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="top", indent=2, wrap_text=True)
        c.border    = _grid
        ws_settings.row_dimensions[r].height = height
        return r

    def _spacer():
        r = _next()
        for ci in (COL_A, COL_B, COL_C):
            c = ws_settings.cell(r, ci, "")
            c.fill   = PatternFill("solid", fgColor=_ROW_ODD)
            c.border = _grid
        ws_settings.row_dimensions[r].height = 8
        return r

    # ── Get risk ranges ──────────────────────────────────────────────────
    _risk_ranges = getattr(self_app, "_sim_risk_ranges", None) or {}
    _low_r  = _risk_ranges.get("LOW",    (1.0,  35.0))
    _med_r  = _risk_ranges.get("MEDIUM", (36.0, 70.0))
    _high_r = _risk_ranges.get("HIGH",   (70.01, float("inf")))

    def _fmt_range(lo, hi):
        return f"{lo:.0f}% – {hi:.0f}%" if hi != float("inf") else f">= {lo:.0f}%"

    # ══════════════════════════════════════════════════════════════════
    #  SECTION 1 — Formula Explanation
    # ══════════════════════════════════════════════════════════════════
    _banner("How % Net Income to Amortization is Calculated")
    _expl_row(
        "The % Net Income to Amortization measures how much of a client's simulated net income "
        "is consumed by their loan amortization obligations. A higher percentage means a larger "
        "portion of income is used for debt repayment, leaving less buffer for operating costs "
        "and unexpected expenses.",
        height=36,
    )
    _sub_header("Step", "Formula / Value", "Description")
    _row3("1. Base Net Income",         " = Total Net Income (Base)",
          "The client's net income before any simulated cost shocks.", height=18)
    _row3("2. Simulated Cost Increase",  " = Sum of (Base Expense × Slider %)",
          "For each modified expense, the extra cost is: Base Amount × (Inflation Rate / 100). "
          "The total of all extra costs is the Simulated Increase.", bg=_ROW_EVEN, height=24, wrap_c=True)
    _row3("3. Simulated Net Income",     " = Base Net Income − Simulated Cost Increase",
          "Net income after the simulated cost shock is applied. This is the denominator "
          "used in the ratio.", height=24, wrap_c=True)
    _row3("4. % Net → Amort",
          " = (Total Current Amort / Simulated Net Income) × 100",
          "The final ratio. If Simulated Net Income ≤ 0, the ratio is set to 999% to flag "
          "the client as extreme risk (amortization exceeds income).",
          bg=_ROW_EVEN, a_bold=True, b_bold=True, height=30, wrap_c=True)

    _spacer()

    # ── IMPORTANT NOTICE — Amortization assumption disclaimer ────────
    _notice_row = _next()
    ws_settings.merge_cells(start_row=_notice_row, start_column=COL_A, end_row=_notice_row, end_column=COL_C)
    _nc = ws_settings.cell(_notice_row, COL_A,
        "⚠  IMPORTANT NOTICE: Not all Net Income and Amortization figures shown in this report are "
        "guaranteed to be accurate. The simulator ASSUMES that ALL clients have an amortization "
        "obligation. This may not reflect reality — some clients may have NO active amortization. "
        "Please double-check each client record to verify whether the client actually has "
        "amortization before acting on any risk label or percentage shown here."
    )
    _nc.fill      = PatternFill("solid", fgColor="FFF3CD")
    _nc.font      = Font(bold=True, size=9, color="856404", name="Calibri")
    _nc.alignment = Alignment(horizontal="left", vertical="top", indent=2, wrap_text=True)
    _nc.border    = Border(
        left=Side(border_style="medium", color="856404"),
        right=Side(border_style="medium", color="856404"),
        top=Side(border_style="medium", color="856404"),
        bottom=Side(border_style="medium", color="856404"),
    )
    ws_settings.row_dimensions[_notice_row].height = 60

    _spacer()

    # ══════════════════════════════════════════════════════════════════
    #  SECTION 2 — Risk Label Explanation
    # ══════════════════════════════════════════════════════════════════
    _banner("How the Risk Label is Determined")
    _expl_row(
        "The Risk Label (LOW / MEDIUM / HIGH) is assigned based on the % Net Income to Amortization "
        "computed after applying the simulated expense increases. The thresholds below are configurable "
        "via the '⚖ Risk Ranges' button in the simulator.",
        height=36,
    )
    _sub_header("Risk Label", "% Net → Amort Range", "Explanation")

    # LOW
    _row3("LOW", _fmt_range(*_low_r),
          f"A client is rated LOW risk when their % Net → Amort falls between "
          f"{_low_r[0]:.0f}% and {_low_r[1]:.0f}%. This means their simulated net income "
          f"comfortably covers their amortization obligation with significant income remaining.",
          a_bold=True, a_fg=_MOD_POS, b_fg=_MOD_POS, height=36, wrap_c=True)

    # MEDIUM
    _row3("MEDIUM", _fmt_range(*_med_r),
          f"A client is rated MEDIUM risk when their % Net → Amort falls between "
          f"{_med_r[0]:.0f}% and {_med_r[1]:.0f}%. The simulated cost shock has notably reduced "
          f"their income buffer, and close monitoring is recommended.",
          bg=_ROW_EVEN, a_bold=True, a_fg=_GOLD, b_fg=_GOLD, height=36, wrap_c=True)

    # HIGH
    _hi_lo = _high_r[0]
    _row3("HIGH", _fmt_range(*_high_r),
          f"A client is rated HIGH risk when their % Net → Amort is {_hi_lo:.0f}% or above "
          f"(including the extreme 999% case where simulated net income is zero or negative). "
          f"This client may have difficulty servicing their loan under the simulated scenario "
          f"and should be reviewed carefully — they might have a special loan case.",
          a_bold=True, a_fg=_MOD_NEG, b_fg=_MOD_NEG, height=48, wrap_c=True)

    _spacer()

    # ══════════════════════════════════════════════════════════════════
    #  SECTION 3 — Expense Modifications
    # ══════════════════════════════════════════════════════════════════
    _banner("Expense Modifications Applied in this Simulation")
    _expl_row(
        "The table below shows every expense item whose inflation slider was set to a non-zero "
        "value. Only modified expenses affect the simulated totals; items at 0% are unchanged. "
        "A positive % means the expense was inflated (cost shock scenario); a negative % "
        "means the expense was reduced.",
        height=36,
    )
    _sub_header("Expense Item", "Change Applied (%)", "Interpretation")

    sliders  = getattr(self_app, "_sim_sliders",  {}) or {}
    expenses = getattr(self_app, "_sim_expenses", []) or []
    modified = []
    zero_list = []
    for exp in expenses:
        name = str((exp or {}).get("name") or "").strip()
        if not name:
            continue
        var = sliders.get(name)
        pct = 0.0
        if var:
            try:
                pct = float(var.get() or 0)
            except Exception:
                pct = 0.0
        if pct != 0.0:
            modified.append((name, pct))
        else:
            zero_list.append(name)
    for nm, var in sliders.items():
        if nm not in {e.get("name", "") for e in expenses}:
            try:
                pct = float(var.get() or 0)
            except Exception:
                pct = 0.0
            if pct != 0.0:
                modified.append((nm, pct))

    if modified:
        modified.sort(key=lambda x: -abs(x[1]))
        for idx, (nm, pct) in enumerate(modified):
            bg = _ROW_ODD if idx % 2 == 0 else _ROW_EVEN
            change_str = f"+{pct:.1f}% increase" if pct > 0 else f"{pct:.1f}% decrease"
            val_color  = _MOD_POS if pct > 0 else _MOD_NEG
            interp = (
                f"Each peso of '{nm}' is multiplied by {1 + pct/100:.3f}x in the simulation. "
                f"Extra cost per client = their '{nm}' base amount × {pct:.1f} / 100."
            )
            _row3(nm, change_str, interp, bg=bg,
                  b_bold=True, b_fg=val_color, height=24, wrap_c=True)
    else:
        r = _next()
        ws_settings.merge_cells(start_row=r, start_column=COL_A, end_row=r, end_column=COL_C)
        c = ws_settings.cell(r, COL_A, "— No expense modifications — all sliders at 0% —")
        c.fill      = PatternFill("solid", fgColor=_ROW_ODD)
        c.font      = Font(italic=True, size=9, color=_NONE_FG, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c.border    = _grid
        ws_settings.row_dimensions[r].height = 16

    # Summary count
    _total_exp = len(modified) + len(zero_list)
    r_sum = _next()
    ws_settings.merge_cells(start_row=r_sum, start_column=COL_A, end_row=r_sum, end_column=COL_C)
    c_sum = ws_settings.cell(r_sum, COL_A,
        f"  {len(modified)} of {_total_exp} expense item(s) modified   |   "
        f"{len(zero_list)} at 0% (no change)"
    )
    c_sum.fill      = PatternFill("solid", fgColor=_HDR_BG)
    c_sum.font      = Font(italic=True, size=8, color=_HDR_FG, name="Calibri")
    c_sum.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c_sum.border    = _grid
    ws_settings.row_dimensions[r_sum].height = 14

    _spacer()

    # ══════════════════════════════════════════════════════════════════
    #  SECTION 4 — Industry Checklist
    # ══════════════════════════════════════════════════════════════════
    _banner("Industry Checklist Used in Simulation")
    _expl_row(
        "Only clients belonging to the industries marked '✔ Yes — included' below were "
        "included in the simulator's Client Impact table. Industries marked '✘ Not selected' "
        "were excluded from the simulation. If no industry filter was active, all industries "
        "are included.",
        height=36,
    )
    _sub_header("Industry", "In Checklist?", "Notes")

    selected_inds = sorted(
        getattr(self_app, "_sim_selected_industries", set()) or set(), key=str.lower
    )
    all_data = getattr(self_app, "_lu_all_data", {}) or {}
    all_inds = sorted(
        {str(x).strip() for x in all_data.get("unique_industries", []) if str(x).strip()},
        key=str.lower,
    )
    selected_set = {str(x).strip().lower() for x in selected_inds}

    # When nothing is checked, the filter is inactive — all industries are included.
    no_filter_active = len(selected_inds) == 0

    if all_inds:
        for idx, ind in enumerate(all_inds):
            bg  = _ROW_ODD if idx % 2 == 0 else _ROW_EVEN
            if no_filter_active:
                # No filter → every industry is included
                status = "✔  Yes — included (no filter active)"
                note   = "All industries included — no industry filter was set."
                b_bold, b_fg, c_fg = True, _MOD_POS, _EXPL_FG
            else:
                chk    = ind.strip().lower() in selected_set
                status = "✔  Yes — included in simulation" if chk else "✘  Not selected"
                note   = ("Clients in this industry were included in the simulator." if chk
                          else "Clients in this industry were excluded from the simulator.")
                b_bold = chk
                b_fg   = _MOD_POS if chk else _NONE_FG
                c_fg   = _EXPL_FG if chk else _NONE_FG
            _row3(ind, status, note, bg=bg,
                  b_bold=b_bold, b_fg=b_fg, c_fg=c_fg,
                  height=16, wrap_c=False)
    elif selected_inds:
        for idx, ind in enumerate(selected_inds):
            bg = _ROW_ODD if idx % 2 == 0 else _ROW_EVEN
            _row3(ind, "✔  Yes — included in simulation",
                  "Clients in this industry were included in the simulator.",
                  bg=bg, b_bold=True, b_fg=_MOD_POS, c_fg=_EXPL_FG, height=16)
    else:
        r = _next()
        ws_settings.merge_cells(start_row=r, start_column=COL_A, end_row=r, end_column=COL_C)
        c = ws_settings.cell(r, COL_A, "— No industry filter active — all industries included —")
        c.fill      = PatternFill("solid", fgColor=_ROW_ODD)
        c.font      = Font(italic=True, size=9, color=_NONE_FG, name="Calibri")
        c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
        c.border    = _grid
        ws_settings.row_dimensions[r].height = 16

    # Summary
    r_ind_sum2 = _next()
    ws_settings.merge_cells(start_row=r_ind_sum2, start_column=COL_A, end_row=r_ind_sum2, end_column=COL_C)
    _summary_text = (
        "  All industries included — no filter active"
        if no_filter_active
        else f"  {len(selected_inds)} of {len(all_inds) or len(selected_inds)} "
             f"industry/industries selected in checklist"
    )
    c_ind2 = ws_settings.cell(r_ind_sum2, COL_A, _summary_text)
    c_ind2.fill      = PatternFill("solid", fgColor=_HDR_BG)
    c_ind2.font      = Font(italic=True, size=8, color=_HDR_FG, name="Calibri")
    c_ind2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    c_ind2.border    = _grid
    ws_settings.row_dimensions[r_ind_sum2].height = 14

    _spacer()


def _sim_export_client_impact_excel(self):
    """
    Export the simulator "Client Impact" rows to Excel.

    - Uses current simulator slider values (what-if).
    - Uses the current LU industry filter already applied to `_sim_recs`.
    - Uses the current client search filter from `_sim_client_search_var`.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messagebox.showerror(
            "Missing Library",
            "openpyxl is not installed.\nRun:  pip install openpyxl",
            parent=self,
        )
        return

    recs_all = list(getattr(self, "_sim_recs", []) or [])
    if not recs_all:
        messagebox.showinfo(
            "No simulator data",
            "Run LU analysis and the simulator first, then try Export again.",
            parent=self,
        )
        return

    # Apply client name search filter (same logic as the on-screen table).
    _search_term = str(
        getattr(self, "_sim_client_search_var", None) and self._sim_client_search_var.get()
        or ""
    ).strip().lower()
    if _search_term:
        recs_all = [
            r for r in recs_all
            if _search_term in str((r or {}).get("client") or "").lower()
        ]
    if not recs_all:
        messagebox.showinfo(
            "No matching rows",
            "No simulator rows match the current client search filter.",
            parent=self,
        )
        return

    # Build per-client simulated totals and risk labels (same rules as the table).
    rows = []
    for rec in recs_all:
        name = str((rec or {}).get("client") or "").strip() or "—"
        industry = str((rec or {}).get("industry") or "").strip() or "—"
        base_net = float((rec or {}).get("net_income") or 0.0)
        current_am = float((rec or {}).get("current_amort") or 0.0)

        base_by_name: dict[str, float] = {}
        base_total_exp = 0.0
        for exp in (rec or {}).get("expenses", []) or []:
            nm = str((exp or {}).get("name") or "").strip()
            if not nm:
                continue
            try:
                amt = float((exp or {}).get("total") or 0.0)
            except Exception:
                amt = 0.0
            if amt <= 0:
                continue
            base_by_name[nm] = base_by_name.get(nm, 0.0) + amt
            base_total_exp += amt

        extra_total = 0.0
        for nm, base_amt in base_by_name.items():
            extra, _sim = _sim_amount_for_expense(self, base_amt, nm)
            extra_total += extra

        sim_total_exp = base_total_exp + extra_total
        pct_inc = (
            extra_total / base_total_exp * 100.0
            if base_total_exp > 0
            else (0.0 if extra_total <= 0 else 100.0)
        )

        # Simulated net income = Base net income minus simulated cost increase.
        sim_net_income = base_net - extra_total
        if current_am == 0:
            pct_net_to_am = -2.0  # sentinel: no amortization — risk is N/A
        elif sim_net_income <= 0:
            pct_net_to_am = -1.0  # sentinel: negative income
        else:
            pct_net_to_am = (
                (current_am / sim_net_income) * 100.0
            )

        sim_risk = _sim_pct_net_to_amort_label(pct_net_to_am, getattr(self, "_sim_risk_ranges", None))
        risk_reasoning = _sim_build_risk_reasoning(sim_risk, pct_net_to_am)
        rows.append({
            "client_id": str((rec or {}).get("client_id") or "").strip(),
            "pn": str((rec or {}).get("pn") or "").strip(),
            "client": name,
            "residence_address": str((rec or {}).get("residence_address") or "").strip(),
            "office_address": str((rec or {}).get("office_address") or "").strip(),
            "industry": industry,
            "loan_status": str((rec or {}).get("loan_status") or "").strip(),
            "ao_name": str((rec or {}).get("ao_name") or "").strip(),
            "ci_bi_date": str((rec or {}).get("ci_bi_date") or "").strip(),
            "product_name": str((rec or {}).get("product_name") or "").strip(),
            "loan_balance": float((rec or {}).get("loan_balance") or 0.0),
            "principal_loan": float((rec or {}).get("principal_loan") or 0.0),
            "base_total_expenses": base_total_exp,
            "sim_total_expenses": sim_total_exp,
            "net_income": base_net,
            "sim_net_income": sim_net_income,
            "pct_increase": pct_inc,
            "sim_increase": extra_total,
            "current_amort": current_am,
            "pct_net_to_amort": pct_net_to_am,
            "sim_risk_label": sim_risk,
            "risk_reasoning": risk_reasoning,
        })

    # Sort by highest % Net → Amort first (descending).
    rows.sort(key=lambda r: -(r["pct_net_to_amort"] or 0.0))

    from tkinter import filedialog
    import getpass
    from datetime import datetime
    from pathlib import Path

    default_name = f"RiskSimulator_ClientImpact_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = filedialog.asksaveasfilename(
        parent=self,
        title="Save Simulator Client Impact Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=default_name,
    )
    if not path:
        return

    wb = openpyxl.Workbook()
    ws_cfg = wb.active
    ws_cfg.title = "Settings"

    exported_by = str(getattr(self, "_current_username", "") or "").strip()
    if not exported_by:
        try:
            exported_by = (getpass.getuser() or "").strip()
        except Exception:
            exported_by = ""
    if not exported_by:
        exported_by = "Unknown user"

    selected_inds = sorted(getattr(self, "_sim_selected_industries", set()) or set(), key=str.lower)
    industry_note = "None (no industry filter)" if not selected_inds else " · ".join(selected_inds)
    search_note = _search_term if _search_term else "None"

    _sim_write_export_settings_sheet(
        ws_cfg,
        self_app=self,
        fname=Path(str(getattr(self, "_lu_filepath", "") or "—")).name,
        generated_at=datetime.now().strftime("%B %d, %Y  %H:%M"),
        exported_by=exported_by,
        export_scope_note=(
            "Risk Simulator export — includes current simulator what-if values from the Client Impact table. "
            f"Client search: {search_note}."
        ),
    )

    # ── Append formula/risk/expense/industry sections to the same sheet ──
    _used_rows = ws_cfg.max_row
    _sim_write_settings_sheet(ws_cfg, self_app=self, row_offset=_used_rows + 1)

    # ── Shared modern style palette ───────────────────────────────────
    _C_HDR_BG    = "1A3A5C"   # deep navy header
    _C_HDR_FG    = "FFFFFF"   # white header text
    _C_BANNER_BG = "1E4D7B"   # slightly lighter navy for banner
    _C_BANNER_FG = "FFFFFF"
    _C_META_BG   = "EAF2FB"   # very light blue for meta row
    _C_META_FG   = "1A3A5C"
    _C_ROW_ODD   = "FFFFFF"   # clean white
    _C_ROW_EVEN  = "F4F8FD"   # subtle blue-tint stripe
    _C_HIGH_BG   = "FEF0EE"   # soft red tint
    _C_HIGH_FG   = "C0392B"
    _C_MED_BG    = "FEFAE8"   # soft amber tint
    _C_MED_FG    = "9A6700"
    _C_LOW_BG    = "EDFAF1"   # soft green tint
    _C_LOW_FG    = "1E8449"
    _C_BORDER    = "D5DCE4"   # light grey grid lines
    _C_ACCENT    = "2980B9"   # blue accent for numbers

    _thin    = Side(border_style="thin",   color=_C_BORDER)
    _thick_b = Side(border_style="medium", color=_C_HDR_BG)
    _grid    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _hdr_bot = Border(left=_thin, right=_thin, top=_thin, bottom=_thick_b)

    NUM_FMT  = '"₱"#,##0.00'
    PCT_FMT  = '0.0"%"'

    # ── Client Impact sheet ───────────────────────────────────────────
    ws = wb.create_sheet("Client Impact")
    NUM_COLS = 22

    # Row 1 — Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    _b = ws.cell(1, 1, "📊  RISK SIMULATOR — Client Impact Report")
    _b.fill      = PatternFill("solid", fgColor=_C_BANNER_BG)
    _b.font      = Font(bold=True, size=13, color=_C_BANNER_FG, name="Calibri")
    _b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Row 2 — Meta info (timestamp / exported by)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    _meta_text = (
        f"Generated: {datetime.now().strftime('%B %d, %Y  %H:%M')}     "
        f"Exported by: {exported_by}     "
        f"Total clients: {len(rows)}"
    )
    _m = ws.cell(2, 1, _meta_text)
    _m.fill      = PatternFill("solid", fgColor=_C_META_BG)
    _m.font      = Font(italic=True, size=9, color=_C_META_FG, name="Calibri")
    _m.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # Row 3 — Column headers
    headers = [
        "Client ID", "PN", "Client", "Residence Address", "Office Address",
        "Industry", "Loan Status", "AO Name", "CI/BI Date", "Product Name",
        "Loan Balance", "Principal Loan",
        "Total Expenses (Base)", "Total Expenses (Sim)",
        "Total Net Income (Base)", "Total Net Income (Sim)",
        "% Increase", "Simulated Increase",
        "Total Current Amort", "% Net → Amort",
        "Risk Label", "Risk Reasoning",
    ]
    _hf = Font(bold=True, size=9, color=_C_HDR_FG, name="Calibri")
    _ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(3, ci, h)
        cell.fill      = PatternFill("solid", fgColor=_C_HDR_BG)
        cell.font      = _hf
        cell.alignment = _ha
        cell.border    = _hdr_bot
    ws.row_dimensions[3].height = 32

    # Column widths
    col_widths = [13, 13, 24, 22, 22, 18, 14, 15, 15, 20, 15, 15, 18, 18, 18, 18, 11, 17, 17, 13, 12, 48]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows (start row 4)
    _risk_colors = {
        "HIGH":                  (_C_HIGH_BG, _C_HIGH_FG),
        "HIGH (NEGATIVE INCOME)": (_C_HIGH_BG, _C_HIGH_FG),
        "MEDIUM":                (_C_MED_BG,  _C_MED_FG),
        "LOW":                   (_C_LOW_BG,  _C_LOW_FG),
    }

    for idx, r in enumerate(rows):
        row_num = 4 + idx
        risk = str(r.get("sim_risk_label") or "").upper()
        if risk in _risk_colors:
            row_bg, _risk_fg = _risk_colors[risk]
        else:
            row_bg = _C_ROW_ODD if idx % 2 == 0 else _C_ROW_EVEN
            _risk_fg = "374151"

        _base_font = Font(size=9, color="374151", name="Calibri")
        _num_font  = Font(size=9, color=_C_ACCENT, name="Calibri")
        _pct_font  = Font(size=9, color="374151", name="Calibri")
        _risk_font = Font(bold=True, size=9, color=_risk_fg, name="Calibri")
        _rf        = PatternFill("solid", fgColor=row_bg)
        _al_l      = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        _al_c      = Alignment(horizontal="center", vertical="center")
        _al_r      = Alignment(horizontal="right",  vertical="center")

        def _wc(col, val, font=_base_font, align=_al_l, fmt=None):
            c = ws.cell(row_num, col, val)
            c.fill = _rf; c.font = font; c.alignment = align; c.border = _grid
            if fmt: c.number_format = fmt
            return c

        _wc(1,  r["client_id"] or "—")
        _wc(2,  r["pn"] or "—")
        _wc(3,  r["client"])
        _wc(4,  r["residence_address"] or "—")
        _wc(5,  r["office_address"] or "—")
        _wc(6,  r["industry"])
        _wc(7,  r["loan_status"] or "—", align=_al_c)
        _wc(8,  r["ao_name"] or "—")
        _wc(9,  r["ci_bi_date"] or "—")
        _wc(10, r["product_name"] or "—")
        _wc(11, r["loan_balance"],        font=_num_font, align=_al_r, fmt=NUM_FMT)
        _wc(12, r["principal_loan"],      font=_num_font, align=_al_r, fmt=NUM_FMT)
        _wc(13, r["base_total_expenses"], font=_num_font, align=_al_r, fmt=NUM_FMT)
        _wc(14, r["sim_total_expenses"],  font=_num_font, align=_al_r, fmt=NUM_FMT)
        _wc(15, r["net_income"],          font=_num_font, align=_al_r, fmt=NUM_FMT)
        _wc(16, r["sim_net_income"],      font=_num_font, align=_al_r, fmt=NUM_FMT)
        _wc(17, r["pct_increase"],        font=_pct_font, align=_al_c, fmt=PCT_FMT)
        _wc(18, r["sim_increase"],        font=_num_font, align=_al_r, fmt=NUM_FMT)
        _wc(19, r["current_amort"],       font=_num_font, align=_al_r, fmt=NUM_FMT)
        _wc(20, r["pct_net_to_amort"],    font=_pct_font, align=_al_c, fmt=PCT_FMT)
        _wc(21, r["sim_risk_label"],      font=_risk_font, align=_al_c)
        _wc(22, r["risk_reasoning"],      align=Alignment(horizontal="left", vertical="center", wrap_text=True))
        ws.row_dimensions[row_num].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(NUM_COLS)}{3 + len(rows)}"
    wb.save(path)
    messagebox.showinfo("Export Complete", f"Excel saved to:\n{path}", parent=self)


def _sim_merge_excel_files(self):
    """
    Merge sample Excel file columns into the live Client Impact table.

    Workflow:
      1. User selects one or more Excel/CSV files (e.g. MARCH_BORROWERS_LIST.xlsx).
      2. ALL columns from those files are read.
      3. Each sample-file row is matched to a Client Impact row by NAME
         (sample "Applicant" ↔ simulator "client").
      4. The output contains EVERY column from both the Client Impact table
         AND the sample file(s), joined on name — all columns merged.

    Merge key is NAME-BASED.  "LAST, FIRST" form is normalised to
    "FIRST LAST" before comparison so name formats don't cause mismatches.
    When a name appears in both sources, sample-file values fill any
    blank Client Impact fields; neither source overwrites existing data.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messagebox.showerror(
            "Missing Library",
            "openpyxl is not installed.\nRun:  pip install openpyxl",
            parent=self,
        )
        return

    # ── Guard: need live simulator data ──────────────────────────────
    recs_all = list(getattr(self, "_sim_recs", []) or [])
    if not recs_all:
        messagebox.showinfo(
            "No Simulator Data",
            "Run LU analysis and the simulator first, then use Merge Excel.",
            parent=self,
        )
        return

    # ── Pick sample file(s) ──────────────────────────────────────────
    paths = filedialog.askopenfilenames(
        parent=self,
        title="Select Excel/CSV file(s) to merge into Client Impact",
        filetypes=[
            ("Excel & CSV files", "*.xlsx *.csv"),
            ("Excel files", "*.xlsx"),
            ("CSV files", "*.csv"),
            ("All files", "*.*"),
        ],
    )
    if not paths:
        return

    # ── Helpers ───────────────────────────────────────────────────────
    def _norm_text(v) -> str:
        return str(v or "").strip()

    def _norm_name(name: str) -> str:
        """Normalise to uppercase, collapse spaces, flip LAST, FIRST → FIRST LAST."""
        s = _norm_text(name).upper()
        s = re.sub(r"[^A-Z0-9\s,.\-]", " ", s)
        s = re.sub(r"\s+", " ", s).strip()
        if "," in s:
            parts = [p.strip() for p in s.split(",", 1)]
            if len(parts) == 2 and parts[0] and parts[1]:
                s = f"{parts[1]} {parts[0]}".strip()
        return s

    def _sample_row_name(row: dict) -> str:
        """Extract the name field from a sample-file row."""
        for col in ("Applicant", "Client", "Client Name", "Name"):
            v = _norm_text(row.get(col))
            if v:
                return v
        return ""

    def _is_empty(v) -> bool:
        if v is None:
            return True
        if isinstance(v, str):
            return v.strip() in ("", "—", "-", "N/A", "n/a")
        return False

    def _read_rows_from_file(path: str) -> tuple[list[str], list[dict]]:
        """Return (headers, rows) from an xlsx or csv file."""
        if str(path).lower().endswith(".csv"):
            import csv as _csv
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = _csv.DictReader(f)
                headers = [str(h or "").strip() for h in (reader.fieldnames or [])]
                rows = []
                for r in reader:
                    rr = {str(k or "").strip(): ("" if v is None else str(v).strip())
                          for k, v in (r or {}).items()}
                    if any(str(v).strip() for v in rr.values()):
                        rows.append(rr)
                return headers, rows

        wb_s = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            ws_s = wb_s.active
            all_rows = list(ws_s.iter_rows(values_only=True))
            if not all_rows:
                return [], []
            headers = [str(c or "").strip() for c in all_rows[0]]
            rows = []
            for vals in all_rows[1:]:
                if vals is None:
                    continue
                row_dict = {}
                for idx, h in enumerate(headers):
                    if not h:
                        continue
                    vv = vals[idx] if idx < len(vals) else ""
                    row_dict[h] = "" if vv is None else str(vv).strip()
                if any(str(v).strip() for v in row_dict.values()):
                    rows.append(row_dict)
            return headers, rows
        finally:
            wb_s.close()

    # ── Build Client Impact rows from live simulator data ─────────────
    # (same logic as _sim_export_client_impact_excel)
    def _money_str(v) -> str:
        try:
            return f"₱{float(v):,.2f}"
        except Exception:
            return "—"

    def _pct_str(v) -> str:
        try:
            fv = float(v)
            if fv < 0:
                return _NEGATIVE_INCOME_PCT_DISPLAY
            return f"{fv:.1f}%"
        except Exception:
            return "—"

    impact_rows: list[dict] = []
    for rec in recs_all:
        name = str((rec or {}).get("client") or "").strip() or "—"
        industry = str((rec or {}).get("industry") or "").strip() or "—"
        base_net = float((rec or {}).get("net_income") or 0.0)
        current_am = float((rec or {}).get("current_amort") or 0.0)

        base_by_name: dict[str, float] = {}
        base_total_exp = 0.0
        for exp in (rec or {}).get("expenses", []) or []:
            nm = str((exp or {}).get("name") or "").strip()
            if not nm:
                continue
            try:
                amt = float((exp or {}).get("total") or 0.0)
            except Exception:
                amt = 0.0
            if amt <= 0:
                continue
            base_by_name[nm] = base_by_name.get(nm, 0.0) + amt
            base_total_exp += amt

        extra_total = 0.0
        for nm, base_amt in base_by_name.items():
            extra, _sim_amt = _sim_amount_for_expense(self, base_amt, nm)
            extra_total += extra

        sim_total_exp = base_total_exp + extra_total
        pct_inc = (
            extra_total / base_total_exp * 100.0
            if base_total_exp > 0
            else (0.0 if extra_total <= 0 else 100.0)
        )
        sim_net_income = base_net - extra_total
        if current_am == 0:
            pct_net_to_am = -2.0  # sentinel: no amortization — risk is N/A
        elif sim_net_income <= 0:
            pct_net_to_am = -1.0  # sentinel: negative income
        else:
            pct_net_to_am = (current_am / sim_net_income) * 100.0

        sim_risk = _sim_pct_net_to_amort_label(pct_net_to_am, getattr(self, "_sim_risk_ranges", None))
        risk_reasoning = _sim_build_risk_reasoning(sim_risk, pct_net_to_am)

        impact_rows.append({
            # Client Impact columns (using the same header names as the export)
            "Client":                   name,
            "Industry":                 industry,
            "Client ID":                str((rec or {}).get("client_id") or "").strip(),
            "PN":                       str((rec or {}).get("pn") or "").strip(),
            "Residence Address":        str((rec or {}).get("residence_address") or "").strip(),
            "Office Address":           str((rec or {}).get("office_address") or "").strip(),
            "Loan Status":              str((rec or {}).get("loan_status") or "").strip(),
            "AO Name":                  str((rec or {}).get("ao_name") or "").strip(),
            "CI/BI Date":               str((rec or {}).get("ci_bi_date") or "").strip(),
            "Product Name":             str((rec or {}).get("product_name") or "").strip(),
            "Loan Balance":             _money_str((rec or {}).get("loan_balance")),
            "Principal Loan":           _money_str((rec or {}).get("principal_loan")),
            "Total Expenses (Base)":    _money_str(base_total_exp),
            "Total Expenses (Sim)":     _money_str(sim_total_exp),
            "Total Net Income (Base)":  _money_str(base_net),
            "Total Net Income (Sim)":   _money_str(sim_net_income),
            "% Increase":               _pct_str(pct_inc),
            "Simulated Increase":       _money_str(extra_total),
            "Total Current Amort":      _money_str(current_am),
            "% Net → Amort":            _pct_str(pct_net_to_am),
            "Risk Label":               sim_risk,
            "Risk Reasoning":           risk_reasoning,
            # Raw floats — used by NIA pre-computation to avoid re-parsing formatted strings
            "_raw_sim_net_income":      sim_net_income,
            "_raw_current_amort":       current_am,
        })

    # Index Client Impact rows by normalised name for O(1) lookup.
    impact_by_norm_name: dict[str, dict] = {}
    for row in impact_rows:
        key = _norm_name(row.get("Client", ""))
        if key:
            impact_by_norm_name[key] = row

    # ── Read all sample files and collect every column ────────────────
    sample_headers: list[str] = []
    sample_by_norm_name: dict[str, dict] = {}
    loaded_rows = 0

    for p in paths:
        try:
            file_headers, file_rows = _read_rows_from_file(p)
        except Exception as exc:
            messagebox.showerror("Merge Error", f"Failed reading:\n{p}\n\n{exc}", parent=self)
            return

        for h in file_headers:
            if h and h not in sample_headers:
                sample_headers.append(h)

        for row in file_rows:
            loaded_rows += 1
            raw_name = _sample_row_name(row)
            key = _norm_name(raw_name) if raw_name else f"__NONAME__{loaded_rows}"
            existing = sample_by_norm_name.get(key)
            if existing is None:
                sample_by_norm_name[key] = dict(row)
            else:
                # Accumulate columns; don't overwrite non-empty values.
                for col, incoming in row.items():
                    if col not in existing:
                        existing[col] = incoming
                    elif _is_empty(existing.get(col)) and not _is_empty(incoming):
                        existing[col] = incoming

    if not sample_by_norm_name:
        messagebox.showinfo("Merge Excel", "No rows found in the selected file(s).", parent=self)
        return

    # ── Build merged output: exact requested column sequence ──────────
    # Priority columns in the exact order the user specified.
    # Columns from the sample file that match priority names are placed
    # in their designated position; any remaining sample columns that
    # don't appear in the priority list are appended at the end.
    PRIORITY_COLS = [
        "Client ID",
        "PN",
        "Applicant",           # maps from sample "Applicant" / simulator "Client"
        "Industry",
        "Total Expenses (Base)",
        "Total Expenses (Sim)",
        "Total Net Income (Base)",
        "Total Net Income (Sim)",
        "% Increase",
        "Simulated Increase",
        "Total Current Amortization",
        "% Net → Amort",
        "Loan Balance",
        "Principal Loan",
        "Maturity",
        "Interest Rate",
        "Branch",
        "Loan Class",
        "Product Name",
        "Loan Date",
        "Term Unit",
        "Term",
        "Security",
        "Release Tag",
        "Loan Amount",
        "Loan Status",
        "Risk Label",
        "Risk Reasoning",
        "AO Name",
        "CI/BI Date",
    ]

    # Internal impact columns that must be remapped to the priority names above.
    # Old header → new header in PRIORITY_COLS
    _COL_REMAP = {
        "Client":                  "Applicant",
        "Total Current Amort":     "Total Current Amortization",
    }

    # Re-key impact_rows so they use the remapped header names.
    for _r in impact_rows:
        for _old, _new in _COL_REMAP.items():
            if _old in _r:
                _r[_new] = _r.pop(_old)

    # Re-key the lookup index (Client → Applicant in merged rows).
    # (impact_by_norm_name keys are untouched — they are derived from the
    #  original "Client" value before the remap loop above.)

    # Build the set of all headers already covered by PRIORITY_COLS.
    priority_set = set(PRIORITY_COLS)

    # Remove sample-file name columns that duplicate "Applicant" / "Client".
    _name_synonyms = {"Client", "Client Name", "Name"}

    # Any sample-file headers not already in PRIORITY_COLS get appended after.
    extra_sample_cols = [
        h for h in sample_headers
        if h not in priority_set and h not in _name_synonyms
    ]

    # Final ordered header list: priority columns first, then extras.
    all_output_headers = PRIORITY_COLS + extra_sample_cols

    # Patch impact_col_order so the header-colour logic below still works.
    impact_col_order = PRIORITY_COLS

    # ── Update impact_by_norm_name to use "Applicant" after remap ────
    # impact_rows were re-keyed above; rebuild the lookup.
    impact_by_norm_name_remapped: dict[str, dict] = {}
    for row in impact_rows:
        key = _norm_name(row.get("Applicant", ""))
        if key:
            impact_by_norm_name_remapped[key] = row

    # Also update _name_synonyms to exclude "Applicant" from being a dup.
    _name_synonyms_for_merge = {"Client", "Client Name", "Name"}

    # Merge: iterate over Client Impact rows and enrich with sample data.
    # Any sample rows with NO matching Client Impact row are appended at end.
    merged_rows: list[dict] = []
    matched_keys: set[str] = set()

    for impact_row in impact_rows:
        key = _norm_name(impact_row.get("Applicant", ""))
        merged = dict(impact_row)  # start with all Client Impact columns

        sample_row = sample_by_norm_name.get(key)
        if sample_row:
            matched_keys.add(key)
            # Merge ALL sample columns into this row.
            for col, val in sample_row.items():
                # Skip name synonyms (already covered by "Applicant").
                if col in _name_synonyms_for_merge:
                    continue
                # Map "Applicant" from sample to our "Applicant" column.
                dest_col = col
                if col == "Applicant":
                    dest_col = "Applicant"
                if dest_col not in merged:
                    merged[dest_col] = val
                elif _is_empty(merged.get(dest_col)) and not _is_empty(val):
                    merged[dest_col] = val

        merged_rows.append(merged)

    # Append unmatched sample rows (names not found in Client Impact).
    unmatched = 0
    for key, sample_row in sample_by_norm_name.items():
        if key in matched_keys or key.startswith("__NONAME__"):
            continue
        raw_name = _sample_row_name(sample_row)
        row_out = {"Applicant": raw_name}
        for col, val in sample_row.items():
            if col in _name_synonyms_for_merge:
                continue
            if col not in row_out:
                row_out[col] = val
        merged_rows.append(row_out)
        unmatched += 1

    # Sort by highest % Net → Amort first (descending).
    # Guard against "N/A (Negative Income)" or any non-numeric string — treat as infinity so
    # negative-income clients always sort to the top.
    def _sort_pct(r):
        raw = str(r.get("% Net → Amort") or "0").replace("%", "").strip()
        try:
            return -float(raw or 0)
        except (ValueError, TypeError):
            return -float("inf")   # negative income → sort first

    merged_rows.sort(key=_sort_pct)

    # ── Save ─────────────────────────────────────────────────────────
    out_path = filedialog.asksaveasfilename(
        parent=self,
        title="Save Merged Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile="ClientImpact_Merged.xlsx",
    )
    if not out_path:
        return

    wb = openpyxl.Workbook()
    ws_cfg_merge = wb.create_sheet("Settings")
    # Move the active (default) sheet to be the data sheet
    ws = wb.active
    ws.title = "Client Impact Merged"

    # ── Write combined Settings sheet (export info + formula/risk/expense/industry) ──
    import getpass as _gp
    from datetime import datetime as _dt2
    _exp_by = str(getattr(self, "_current_username", "") or "").strip()
    if not _exp_by:
        try: _exp_by = (_gp.getuser() or "").strip()
        except Exception: _exp_by = ""
    if not _exp_by:
        _exp_by = "Unknown user"
    from pathlib import Path as _Path2
    _sim_write_export_settings_sheet(
        ws_cfg_merge,
        self_app=self,
        fname=_Path2(str(getattr(self, "_lu_filepath", "") or "—")).name,
        generated_at=_dt2.now().strftime("%B %d, %Y  %H:%M"),
        exported_by=_exp_by,
        export_scope_note=(
            "Merged Excel export — Client Impact simulator data merged with external file(s). "
            f"Files merged: {len(paths)}."
        ),
    )
    _used_rows_merge = ws_cfg_merge.max_row
    _sim_write_settings_sheet(ws_cfg_merge, self_app=self, row_offset=_used_rows_merge + 1)

    # ── Modern style palette (teal/navy theme) ────────────────────────
    _C_HDR_BG   = "1A3A5C"   # deep navy
    _C_HDR_FG   = "FFFFFF"
    _C_BAN_BG   = "14526E"   # teal-navy banner
    _C_META_BG  = "EAF4F8"
    _C_META_FG  = "14526E"
    _C_ROW_ODD  = "FFFFFF"
    _C_ROW_EVEN = "F2F8FC"   # subtle teal stripe
    _C_HIGH_BG  = "FEF0EE"
    _C_HIGH_FG  = "C0392B"
    _C_MED_BG   = "FEFAE8"
    _C_MED_FG   = "9A6700"
    _C_LOW_BG   = "EDFAF1"
    _C_LOW_FG   = "1E8449"
    _C_NUM      = "1A6E8C"   # teal accent for numeric cells
    _C_BORDER   = "C8D9E4"

    _thin_m  = Side(border_style="thin",   color=_C_BORDER)
    _thick_m = Side(border_style="medium", color=_C_HDR_BG)
    _grid_m  = Border(left=_thin_m, right=_thin_m, top=_thin_m, bottom=_thin_m)
    _hdr_bot = Border(left=_thin_m, right=_thin_m, top=_thin_m, bottom=_thick_m)

    from datetime import datetime as _dt
    _now_str = _dt.now().strftime("%B %d, %Y  %H:%M")
    _total_cols = len(all_output_headers)

    # Row 1 — Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_total_cols)
    _b = ws.cell(1, 1, "🧩  CLIENT IMPACT — Merged Report")
    _b.fill      = PatternFill("solid", fgColor=_C_BAN_BG)
    _b.font      = Font(bold=True, size=13, color=_C_HDR_FG, name="Calibri")
    _b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Row 2 — Meta info
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_total_cols)
    _m2 = ws.cell(2, 1,
        f"Generated: {_now_str}     "
        f"Client Impact rows: {len(impact_rows)}     "
        f"Total columns: {_total_cols}"
    )
    _m2.fill      = PatternFill("solid", fgColor=_C_META_BG)
    _m2.font      = Font(italic=True, size=9, color=_C_META_FG, name="Calibri")
    _m2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # Row 3 — Column headers
    _hf = Font(bold=True, size=9, color=_C_HDR_FG, name="Calibri")
    _ha = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(all_output_headers, 1):
        cell = ws.cell(3, ci, h)
        cell.fill      = PatternFill("solid", fgColor=_C_HDR_BG)
        cell.font      = _hf
        cell.alignment = _ha
        cell.border    = _hdr_bot
        ws.column_dimensions[get_column_letter(ci)].width = 22
    ws.row_dimensions[3].height = 32

    # Numeric / percentage column name sets for smart formatting
    _NUM_COLS = {
        "Total Expenses (Base)", "Total Expenses (Sim)",
        "Total Net Income (Base)", "Total Net Income (Sim)",
        "Simulated Increase", "Total Current Amortization",
        "Loan Balance", "Principal Loan", "Loan Amount",
    }
    _PCT_COLS = {"% Increase", "% Net → Amort", "Interest Rate"}
    NUM_FMT = '"₱"#,##0.00'
    PCT_FMT = '0.0"%"'

    _risk_colors = {
        "HIGH":                  (_C_HIGH_BG, _C_HIGH_FG, True),
        "HIGH (NEGATIVE INCOME)": (_C_HIGH_BG, _C_HIGH_FG, True),
        "MEDIUM":                (_C_MED_BG,  _C_MED_FG,  True),
        "LOW":                   (_C_LOW_BG,  _C_LOW_FG,  True),
    }

    # Data rows (start row 4)
    for ri, row in enumerate(merged_rows):
        row_num = 4 + ri
        risk = str(row.get("Risk Label") or "").upper()
        if risk in _risk_colors:
            row_bg, _risk_fg, _bold_risk = _risk_colors[risk]
        else:
            row_bg   = _C_ROW_ODD if ri % 2 == 0 else _C_ROW_EVEN
            _risk_fg = "374151"
            _bold_risk = False

        _rf        = PatternFill("solid", fgColor=row_bg)
        _base_font = Font(size=9, color="374151", name="Calibri")
        _num_font  = Font(size=9, color=_C_NUM,   name="Calibri")
        _risk_font = Font(bold=True, size=9, color=_risk_fg, name="Calibri")
        _al_l = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        _al_c = Alignment(horizontal="center", vertical="center")
        _al_r = Alignment(horizontal="right",  vertical="center")

        for ci, h in enumerate(all_output_headers, 1):
            raw_val = row.get(h, "")
            # Try to cast numeric/pct columns to float for proper Excel formatting
            if h in _NUM_COLS:
                try:
                    cell_val = float(str(raw_val).replace("₱", "").replace(",", "").strip())
                    cell_fmt = NUM_FMT
                    cell_font = _num_font
                    cell_align = _al_r
                except Exception:
                    cell_val = raw_val; cell_fmt = None
                    cell_font = _base_font; cell_align = _al_l
            elif h in _PCT_COLS:
                try:
                    cell_val = float(str(raw_val).replace("%", "").strip())
                    cell_fmt = PCT_FMT
                    cell_font = _base_font
                    cell_align = _al_c
                except Exception:
                    cell_val = raw_val; cell_fmt = None
                    cell_font = _base_font; cell_align = _al_c
            elif h == "Risk Label":
                cell_val = raw_val; cell_fmt = None
                cell_font = _risk_font; cell_align = _al_c
            else:
                cell_val = raw_val; cell_fmt = None
                cell_font = _base_font; cell_align = _al_l

            c = ws.cell(row_num, ci, cell_val)
            c.fill = _rf; c.font = cell_font
            c.alignment = cell_align; c.border = _grid_m
            if cell_fmt:
                c.number_format = cell_fmt

        ws.row_dimensions[row_num].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(_total_cols)}{3 + len(merged_rows)}"

    # ── NEW SHEET: Net Income + Amortization ─────────────────────────
    ws_nia = wb.create_sheet("Net Income + Amortization")

    _NIA_COLS = [
        "Client ID",
        "Applicant",
        "Net Income + Amortization",
        "% New Net Income to Amortization",
        "Risk Label",
        "Risk Reasoning",
    ]
    _NIA_COL_WIDTHS = [18, 30, 28, 32, 14, 40]
    _nia_total_cols = len(_NIA_COLS)

    # Row 1 — Title banner
    ws_nia.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_nia_total_cols)
    _nia_b = ws_nia.cell(1, 1, "📊  NET INCOME + AMORTIZATION — Analysis Report")
    _nia_b.fill      = PatternFill("solid", fgColor=_C_BAN_BG)
    _nia_b.font      = Font(bold=True, size=13, color=_C_HDR_FG, name="Calibri")
    _nia_b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_nia.row_dimensions[1].height = 28

    # Row 2 — Meta info
    ws_nia.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_nia_total_cols)
    _nia_m2 = ws_nia.cell(2, 1,
        f"Generated: {_now_str}     "
        f"Rows: {len(merged_rows)}     "
        f"Net Income + Amortization = Total Net Income (Sim) + Total Current Amortization"
    )
    _nia_m2.fill      = PatternFill("solid", fgColor=_C_META_BG)
    _nia_m2.font      = Font(italic=True, size=9, color=_C_META_FG, name="Calibri")
    _nia_m2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_nia.row_dimensions[2].height = 16

    # Row 3 — Important Notice
    ws_nia.merge_cells(start_row=3, start_column=1, end_row=3, end_column=_nia_total_cols)
    _nia_notice = ws_nia.cell(3, 1,
        "⚠  IMPORTANT NOTICE: Not all Net Income and Amortization figures in this report are "
        "guaranteed to be accurate. The simulator ASSUMES that ALL clients have an amortization "
        "obligation. This may not reflect reality — some clients may have NO active amortization. "
        "Please DOUBLE-CHECK each client record to verify whether the client actually has "
        "amortization before acting on any risk label or percentage shown in this sheet."
    )
    _nia_notice.fill      = PatternFill("solid", fgColor="FFF3CD")
    _nia_notice.font      = Font(bold=True, size=9, color="856404", name="Calibri")
    _nia_notice.alignment = Alignment(horizontal="left", vertical="top", indent=2, wrap_text=True)
    _nia_notice.border    = Border(
        left=Side(border_style="medium", color="856404"),
        right=Side(border_style="medium", color="856404"),
        top=Side(border_style="medium", color="856404"),
        bottom=Side(border_style="medium", color="856404"),
    )
    ws_nia.row_dimensions[3].height = 54

    # Row 4 — Column headers
    for ci, (h, cw) in enumerate(zip(_NIA_COLS, _NIA_COL_WIDTHS), 1):
        cell = ws_nia.cell(4, ci, h)
        cell.fill      = PatternFill("solid", fgColor=_C_HDR_BG)
        cell.font      = Font(bold=True, size=9, color=_C_HDR_FG, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _hdr_bot
        ws_nia.column_dimensions[get_column_letter(ci)].width = cw
    ws_nia.row_dimensions[4].height = 32

    # Find column indices in "Client Impact Merged" sheet for cross-sheet formulas
    # Columns in all_output_headers that we need to reference:
    def _col_idx(header_name):
        try:
            return all_output_headers.index(header_name) + 1
        except ValueError:
            return None

    _ci_client_id     = _col_idx("Client ID")
    _ci_applicant     = _col_idx("Applicant")
    _ci_sim_net       = _col_idx("Total Net Income (Sim)")
    _ci_amort         = _col_idx("Total Current Amortization")
    _ci_risk_label    = _col_idx("Risk Label")
    _ci_risk_reason   = _col_idx("Risk Reasoning")

    _NIA_NUM_FMT = '"\u20b1"#,##0.00'
    _NIA_PCT_FMT = '0.0"%"'

    # Pre-compute per-row NIA values so Risk Label uses the NIA percentage,
    # not the CIM percentage — these are different formulas and produce
    # different numbers, so the risk label must be evaluated independently.
    _active_ranges = getattr(self, "_sim_risk_ranges", None)

    def _parse_money(val):
        """Robustly parse a formatted money string or raw float back to float.
        Handles: ₱1,234.56  /  ₱-1,234.56  /  -₱1,234.56  /  (₱1,234.56)  / raw float."""
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        # Raw-float key stored by impact_rows builder — use directly
        negative = False
        if s.startswith("(") and s.endswith(")"):
            s = s[1:-1]; negative = True   # accounting notation
        s = s.replace("\u20b1", "").replace("₱", "").replace(",", "").strip()
        if s.startswith("-"):
            negative = not negative   # double-negative → positive
            s = s[1:]
        try:
            v = float(s or 0)
            return -v if negative else v
        except (ValueError, TypeError):
            return 0.0

    _nia_row_meta: list[dict] = []
    for _row in merged_rows:
        # Prefer raw float stored under _raw_ key (written by impact_rows builder);
        # fall back to parsing the formatted string.
        _sn = _row.get("_raw_sim_net_income")
        if _sn is None:
            _sn = _parse_money(_row.get("Total Net Income (Sim)"))
        else:
            _sn = float(_sn)
        _am = _row.get("_raw_current_amort")
        if _am is None:
            _am = _parse_money(_row.get("Total Current Amortization") or _row.get("Total Current Amort"))
        else:
            _am = float(_am)
        _nia_net = _sn + _am
        # Zero amortization → risk is N/A (sentinel -2.0).
        # Negative nia_net → negative income (sentinel -1.0).
        # Otherwise calculate normally.
        if _am == 0:
            _nia_pct = -2.0  # sentinel: no amortization — risk is N/A
        elif _nia_net < 0:
            _nia_pct = -1.0  # sentinel: negative income
        elif _nia_net > 0:
            _nia_pct = _am / _nia_net * 100.0
        else:
            _nia_pct = 0.0
        _nia_risk = _sim_pct_net_to_amort_label(_nia_pct, _active_ranges)
        _nia_reasoning = _sim_build_risk_reasoning(_nia_risk, _nia_pct)
        _nia_row_meta.append({
            "sim_net":       _sn,
            "amort":         _am,
            "nia_net":       _nia_net,
            "nia_pct":       _nia_pct,
            "nia_risk":      _nia_risk,
            "nia_reasoning": _nia_reasoning,
        })

    # Data rows — NIA values are static (computed above); only Client ID / Applicant
    # pull via cross-sheet formula from "Client Impact Merged".
    # NOTE: row_num starts at 5 because row 3 is now the important notice and row 4 is headers.
    for ri, _row in enumerate(merged_rows):
        row_num   = 5 + ri
        src_row   = 4 + ri   # "Client Impact Merged" data still starts at row 4
        src_sheet = "'Client Impact Merged'"

        _meta = _nia_row_meta[ri]
        nia_risk_label = _meta["nia_risk"]
        if nia_risk_label in _risk_colors:
            nia_row_bg, _nia_risk_fg, _ = _risk_colors[nia_risk_label]
        else:
            nia_row_bg   = _C_ROW_ODD if ri % 2 == 0 else _C_ROW_EVEN
            _nia_risk_fg = "374151"

        _nia_rf        = PatternFill("solid", fgColor=nia_row_bg)
        _nia_base_font = Font(size=9, color="374151", name="Calibri")
        _nia_num_font  = Font(size=9, color=_C_NUM,   name="Calibri")
        _nia_risk_font = Font(bold=True, size=9, color=_nia_risk_fg, name="Calibri")
        _nia_pct_font  = Font(size=9, color="374151", name="Calibri")
        _nia_al_l = Alignment(horizontal="left",   vertical="center")
        _nia_al_c = Alignment(horizontal="center", vertical="center")
        _nia_al_r = Alignment(horizontal="right",  vertical="center")

        # Col 1: Client ID
        if _ci_client_id:
            c1 = ws_nia.cell(row_num, 1, f"={src_sheet}!{get_column_letter(_ci_client_id)}{src_row}")
        else:
            c1 = ws_nia.cell(row_num, 1, str(_row.get("Client ID") or ""))
        c1.fill = _nia_rf; c1.font = _nia_base_font
        c1.alignment = _nia_al_l; c1.border = _grid_m

        # Col 2: Applicant
        if _ci_applicant:
            c2 = ws_nia.cell(row_num, 2, f"={src_sheet}!{get_column_letter(_ci_applicant)}{src_row}")
        else:
            c2 = ws_nia.cell(row_num, 2, str(_row.get("Applicant") or ""))
        c2.fill = _nia_rf; c2.font = _nia_base_font
        c2.alignment = _nia_al_l; c2.border = _grid_m

        # Col 3: Net Income + Amortization — static computed value
        c3 = ws_nia.cell(row_num, 3, _meta["nia_net"])
        c3.number_format = _NIA_NUM_FMT
        c3.fill = _nia_rf; c3.font = _nia_num_font
        c3.alignment = _nia_al_r; c3.border = _grid_m

        # Col 4: % New Net Income to Amortization — static computed value,
        # evaluated using the live risk ranges (NOT copied from CIM).
        # Sentinel -2.0 means zero amortization (N/A); -1.0 means negative income.
        _nia_pct_val = _meta["nia_pct"]
        if _nia_pct_val == -2.0:
            c4 = ws_nia.cell(row_num, 4, "N/A (No Amortization)")
            c4.font = Font(italic=True, size=9, color=_C_META_FG, name="Calibri")
        elif _nia_pct_val < 0:
            c4 = ws_nia.cell(row_num, 4, _NEGATIVE_INCOME_PCT_DISPLAY)
            c4.font = Font(italic=True, size=9, color=_C_HIGH_FG, name="Calibri")
        else:
            c4 = ws_nia.cell(row_num, 4, _nia_pct_val)
            c4.number_format = _NIA_PCT_FMT
            c4.font = _nia_pct_font
        c4.fill = _nia_rf
        c4.alignment = _nia_al_c; c4.border = _grid_m

        # Col 5: Risk Label — freshly computed from NIA pct using live risk ranges
        c5 = ws_nia.cell(row_num, 5, nia_risk_label)
        c5.fill = _nia_rf; c5.font = _nia_risk_font
        c5.alignment = _nia_al_c; c5.border = _grid_m

        # Col 6: Risk Reasoning — freshly computed from NIA risk label
        c6 = ws_nia.cell(row_num, 6, _meta["nia_reasoning"])
        c6.fill = _nia_rf; c6.font = _nia_base_font
        c6.alignment = _nia_al_l; c6.border = _grid_m

        ws_nia.row_dimensions[row_num].height = 18

    ws_nia.freeze_panes = "A5"
    ws_nia.auto_filter.ref = f"A4:F{4 + len(merged_rows)}"

    # ── NEW SHEET: Risk Summary ────────────────────────────────────────
    # Shows count of HIGH / MEDIUM / LOW clients based on NIA Risk Label.
    # ─────────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Risk Summary")

    _high_count   = sum(1 for m in _nia_row_meta if m["nia_risk"] == "HIGH" or "NEGATIVE" in str(m.get("nia_risk","")).upper())
    _medium_count = sum(1 for m in _nia_row_meta if m["nia_risk"] == "MEDIUM")
    _low_count    = sum(1 for m in _nia_row_meta if m["nia_risk"] == "LOW")
    _total_count  = len(_nia_row_meta)

    # ── palette (reuse existing vars in scope) ────────────────────────
    _rs_navy     = _C_HDR_BG    # "1A3A5C"
    _rs_white    = "FFFFFF"
    _rs_banner   = _C_BAN_BG    # "14526E"
    _rs_meta_bg  = _C_META_BG   # "EAF4F8"
    _rs_meta_fg  = _C_META_FG   # "14526E"
    _thin_rs  = Side(border_style="thin",   color="D5DCE4")
    _thick_rs = Side(border_style="medium", color=_rs_navy)
    _grid_rs  = Border(left=_thin_rs, right=_thin_rs, top=_thin_rs, bottom=_thin_rs)
    _hdr_rs   = Border(left=_thin_rs, right=_thin_rs, top=_thin_rs, bottom=_thick_rs)

    # Column widths
    ws_sum.column_dimensions["A"].width = 26
    ws_sum.column_dimensions["B"].width = 18
    ws_sum.column_dimensions["C"].width = 18

    # Row 1 — Title banner
    ws_sum.merge_cells("A1:C1")
    _rs_b = ws_sum.cell(1, 1, "📊  RISK SUMMARY — Based on Net Income + Amortization")
    _rs_b.fill      = PatternFill("solid", fgColor=_rs_banner)
    _rs_b.font      = Font(bold=True, size=13, color=_rs_white, name="Calibri")
    _rs_b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_sum.row_dimensions[1].height = 28

    # Row 2 — Meta
    ws_sum.merge_cells("A2:C2")
    _rs_m2 = ws_sum.cell(2, 1,
        f"Generated: {_now_str}     Total clients: {_total_count}     "
        f"Risk thresholds: LOW ≤35% | MEDIUM 36–70% | HIGH >70%"
    )
    _rs_m2.fill      = PatternFill("solid", fgColor=_rs_meta_bg)
    _rs_m2.font      = Font(italic=True, size=9, color=_rs_meta_fg, name="Calibri")
    _rs_m2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_sum.row_dimensions[2].height = 16

    # Row 3 — Column headers
    for ci3, hdr_txt in enumerate(["Risk Level", "Client Count", "% of Total"], 1):
        _hc = ws_sum.cell(3, ci3, hdr_txt)
        _hc.fill      = PatternFill("solid", fgColor=_rs_navy)
        _hc.font      = Font(bold=True, size=9, color=_rs_white, name="Calibri")
        _hc.alignment = Alignment(horizontal="center", vertical="center")
        _hc.border    = _hdr_rs
    ws_sum.row_dimensions[3].height = 28

    # Data rows — HIGH / MEDIUM / LOW
    _sum_data = [
        ("HIGH",   _high_count,   _C_HIGH_BG, _C_HIGH_FG),
        ("MEDIUM", _medium_count, _C_MED_BG,  _C_MED_FG),
        ("LOW",    _low_count,    _C_LOW_BG,  _C_LOW_FG),
    ]
    for row_offset, (lbl, cnt, bg, fg) in enumerate(_sum_data):
        rn = 4 + row_offset
        ws_sum.row_dimensions[rn].height = 24
        pct_of_total = (cnt / _total_count * 100.0) if _total_count > 0 else 0.0

        _ca = ws_sum.cell(rn, 1, lbl)
        _ca.fill      = PatternFill("solid", fgColor=bg)
        _ca.font      = Font(bold=True, size=11, color=fg, name="Calibri")
        _ca.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        _ca.border    = _grid_rs

        _cb = ws_sum.cell(rn, 2, cnt)
        _cb.fill      = PatternFill("solid", fgColor=bg)
        _cb.font      = Font(bold=True, size=11, color=fg, name="Calibri")
        _cb.alignment = Alignment(horizontal="center", vertical="center")
        _cb.border    = _grid_rs

        _cc = ws_sum.cell(rn, 3, f"{pct_of_total:.1f}%")
        _cc.fill      = PatternFill("solid", fgColor=bg)
        _cc.font      = Font(size=10, color=fg, name="Calibri")
        _cc.alignment = Alignment(horizontal="center", vertical="center")
        _cc.border    = _grid_rs

    # Total row
    _rt = 7
    ws_sum.row_dimensions[_rt].height = 22
    for ci3, val in enumerate([f"TOTAL  ({_total_count} clients)", _total_count, "100.0%"], 1):
        _tc = ws_sum.cell(_rt, ci3, val)
        _tc.fill      = PatternFill("solid", fgColor=_rs_navy)
        _tc.font      = Font(bold=True, size=9, color=_rs_white, name="Calibri")
        _tc.alignment = Alignment(horizontal="center" if ci3 > 1 else "left",
                                  vertical="center", indent=(1 if ci3 == 1 else 0))
        _tc.border    = Border(left=_thin_rs, right=_thin_rs,
                               top=_thick_rs, bottom=_thin_rs)

    # ── HIGH RISK CLIENT LIST (below the summary counts) ─────────────
    # Columns: #  |  Client ID  |  Applicant  |  % NIA  |  Net Income+Amort  |  Risk Reasoning
    _rs_hr_cols  = ["#", "Client ID", "Applicant", "% Net → Amort (NIA)", "Net Income + Amort", "Risk Reasoning"]
    _rs_hr_widths = [6, 16, 32, 22, 22, 55]

    # Extend column widths for columns D, E, F (1-indexed: 4,5,6)
    for _ci_ext, _w_ext in enumerate(_rs_hr_widths, 1):
        ws_sum.column_dimensions[get_column_letter(_ci_ext)].width = _w_ext

    # Spacer row
    _rs_spacer = 9
    ws_sum.row_dimensions[8].height = 10

    # Section banner
    ws_sum.merge_cells(f"A{_rs_spacer}:F{_rs_spacer}")
    _rs_hbanner = ws_sum.cell(_rs_spacer, 1, "🔴  HIGH RISK CLIENTS — % Net Income to Amortization > 70%")
    _rs_hbanner.fill      = PatternFill("solid", fgColor=_C_HIGH_FG)
    _rs_hbanner.font      = Font(bold=True, size=11, color="FFFFFF", name="Calibri")
    _rs_hbanner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_sum.row_dimensions[_rs_spacer].height = 24

    # Column headers
    _rs_hr_hdr_row = _rs_spacer + 1
    for _ci_h, _htxt in enumerate(_rs_hr_cols, 1):
        _hcell = ws_sum.cell(_rs_hr_hdr_row, _ci_h, _htxt)
        _hcell.fill      = PatternFill("solid", fgColor=_rs_navy)
        _hcell.font      = Font(bold=True, size=9, color=_rs_white, name="Calibri")
        _hcell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _hcell.border    = _hdr_rs
    ws_sum.row_dimensions[_rs_hr_hdr_row].height = 28

    # Collect HIGH risk rows (parallel zip of merged_rows + _nia_row_meta)
    _high_clients = [
        (mrow, meta)
        for mrow, meta in zip(merged_rows, _nia_row_meta)
        if meta["nia_risk"] == "HIGH" or "NEGATIVE" in str(meta.get("nia_risk","")).upper()
    ]
    # Sort by NIA % descending
    _high_clients.sort(key=lambda x: -(x[1]["nia_pct"] or 0.0))

    if _high_clients:
        _rs_hr_data_start = _rs_hr_hdr_row + 1
        for _hri, (mrow, meta) in enumerate(_high_clients):
            _hrn = _rs_hr_data_start + _hri
            ws_sum.row_dimensions[_hrn].height = 20
            _hr_bg = "FEF0EE"   # soft red tint
            _hr_bg2 = "FDE8E5"  # slightly darker alternating
            _row_bg = _hr_bg if _hri % 2 == 0 else _hr_bg2
            _hr_fill = PatternFill("solid", fgColor=_row_bg)
            _hr_font_lbl = Font(size=9, color="374151", name="Calibri")
            _hr_font_pct = Font(bold=True, size=9, color=_C_HIGH_FG, name="Calibri")

            _cells_vals = [
                (_hri + 1,                     "center", _hr_font_lbl),
                (str(mrow.get("Client ID") or "—"),   "center", _hr_font_lbl),
                (str(mrow.get("Applicant") or "—"),   "left",   _hr_font_lbl),
                ((_NEGATIVE_INCOME_PCT_DISPLAY if meta["nia_pct"] < 0 else f"{meta['nia_pct']:.1f}%"), "center", _hr_font_pct),
                (meta["nia_net"],                       "right",  _hr_font_lbl),
                (meta["nia_reasoning"],                 "left",   _hr_font_lbl),
            ]
            for _ci_d, (_val, _anchor, _fnt) in enumerate(_cells_vals, 1):
                _dc = ws_sum.cell(_hrn, _ci_d, _val)
                _dc.fill      = _hr_fill
                _dc.font      = _fnt
                _dc.alignment = Alignment(horizontal=_anchor, vertical="center",
                                          indent=(1 if _anchor == "left" else 0),
                                          wrap_text=(_ci_d == 6))
                _dc.border    = _grid_rs
                if _ci_d == 5:
                    _dc.number_format = _NIA_NUM_FMT
    else:
        _no_high_row = _rs_hr_hdr_row + 1
        ws_sum.merge_cells(f"A{_no_high_row}:F{_no_high_row}")
        _nc = ws_sum.cell(_no_high_row, 1, "✅  No HIGH risk clients found under current simulator settings.")
        _nc.fill      = PatternFill("solid", fgColor="EDFAF1")
        _nc.font      = Font(italic=True, size=9, color="1E8449", name="Calibri")
        _nc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws_sum.row_dimensions[_no_high_row].height = 20

    # ── NEW SHEET: Sheet Comparison ───────────────────────────────────
    # Compares "Client Impact Merged" vs "Net Income + Amortization" side by side.
    #
    # Columns:
    #   Client ID | Applicant
    #   [CIM] Total Net Income (Sim)  | [NIA] Net Income + Amortization  | Δ Value
    #   [CIM] % Net → Amort           | [NIA] % New Net Income to Amort  | Δ %
    #   [CIM] Risk Label              | [NIA] Risk Label                 | Label Changed?
    # ─────────────────────────────────────────────────────────────────
    ws_cmp = wb.create_sheet("Sheet Comparison")

    _CMP_COLS = [
        "Client ID",
        "Applicant",
        # Net Income comparison
        "Net Income (Sim)\n[Client Impact Merged]",
        "Net Income + Amortization\n[NIA Sheet]",
        "Total Current Amortization\n[Client Impact Merged]",
        # Percentage comparison
        "% Net → Amort\n[Client Impact Merged]",
        "% New Net Income to Amort\n[NIA Sheet]",
        "Δ % (NIA − CIM)",
        # Risk label comparison
        "Risk Label\n[Client Impact Merged]",
        "Risk Label\n[NIA Sheet]",
        "Label Changed?",
    ]
    _CMP_COL_WIDTHS = [16, 28, 28, 28, 26, 24, 26, 20, 20, 20, 16]
    _cmp_total_cols = len(_CMP_COLS)

    # Colour palette — reuse existing vars plus a few comparison-specific ones
    _C_GRP_NET  = "1A5276"   # dark blue  — group header for Net Income columns
    _C_GRP_PCT  = "145A32"   # dark green — group header for % columns
    _C_GRP_RSK  = "6E2F1A"   # dark red   — group header for Risk columns
    _C_DELTA_POS = "E8F8F0"  # light green bg — NIA > CIM
    _C_DELTA_NEG = "FDECEA"  # light red bg   — NIA < CIM
    _C_DELTA_NEU = "F4F6F7"  # neutral grey   — equal
    _C_CHANGED   = "FDECEA"
    _C_SAME      = "EDFAF1"

    _thin_c  = Side(border_style="thin",   color=_C_BORDER)
    _thick_c = Side(border_style="medium", color=_C_HDR_BG)
    _grid_c  = Border(left=_thin_c, right=_thin_c, top=_thin_c, bottom=_thin_c)
    _hdr_c   = Border(left=_thin_c, right=_thin_c, top=_thin_c, bottom=_thick_c)

    # Row 1 — Title banner
    ws_cmp.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_cmp_total_cols)
    _cmp_b = ws_cmp.cell(1, 1, "🔍  SHEET COMPARISON — Client Impact Merged vs Net Income + Amortization")
    _cmp_b.fill      = PatternFill("solid", fgColor=_C_BAN_BG)
    _cmp_b.font      = Font(bold=True, size=13, color=_C_HDR_FG, name="Calibri")
    _cmp_b.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_cmp.row_dimensions[1].height = 28

    # Row 2 — Meta info
    ws_cmp.merge_cells(start_row=2, start_column=1, end_row=2, end_column=_cmp_total_cols)
    _cmp_m2 = ws_cmp.cell(2, 1,
        f"Generated: {_now_str}     "
        f"Rows: {len(merged_rows)}     "
        "Δ = NIA Sheet value minus Client Impact Merged value     "
        "Label Changed? = YES if Risk Label differs between sheets"
    )
    _cmp_m2.fill      = PatternFill("solid", fgColor=_C_META_BG)
    _cmp_m2.font      = Font(italic=True, size=9, color=_C_META_FG, name="Calibri")
    _cmp_m2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws_cmp.row_dimensions[2].height = 16

    # Row 3 — Group sub-headers (colour-coded spans)
    _grp_specs = [
        # (start_col, end_col, label, bg_color)
        (1,  2,  "CLIENT IDENTIFIER",      _C_HDR_BG),
        (3,  5,  "NET INCOME COMPARISON",  _C_GRP_NET),
        (6,  8,  "PERCENTAGE COMPARISON",  _C_GRP_PCT),
        (9,  11, "RISK LABEL COMPARISON",  _C_GRP_RSK),
    ]
    for gc_start, gc_end, g_label, g_bg in _grp_specs:
        if gc_start == gc_end:
            ws_cmp.cell(3, gc_start, g_label)
        else:
            ws_cmp.merge_cells(start_row=3, start_column=gc_start,
                               end_row=3,   end_column=gc_end)
            ws_cmp.cell(3, gc_start, g_label)
        for gci in range(gc_start, gc_end + 1):
            _gc = ws_cmp.cell(3, gci)
            _gc.fill      = PatternFill("solid", fgColor=g_bg)
            _gc.font      = Font(bold=True, size=8, color=_C_HDR_FG, name="Calibri")
            _gc.alignment = Alignment(horizontal="center", vertical="center")
            _gc.border    = _thin_c
    ws_cmp.row_dimensions[3].height = 18

    # Row 4 — Column headers
    for ci, (h, cw) in enumerate(zip(_CMP_COLS, _CMP_COL_WIDTHS), 1):
        cell = ws_cmp.cell(4, ci, h)
        cell.fill      = PatternFill("solid", fgColor=_C_HDR_BG)
        cell.font      = Font(bold=True, size=9, color=_C_HDR_FG, name="Calibri")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _hdr_c
        ws_cmp.column_dimensions[get_column_letter(ci)].width = cw
    ws_cmp.row_dimensions[4].height = 36

    # Helper: column letter lookups in source sheets
    def _cim_col(header_name):
        try:
            return get_column_letter(all_output_headers.index(header_name) + 1)
        except ValueError:
            return None

    _cim_sheet       = "'Client Impact Merged'"
    _nia_sheet       = "'Net Income + Amortization'"

    _cim_client_id   = _cim_col("Client ID")
    _cim_applicant   = _cim_col("Applicant")
    _cim_sim_net     = _cim_col("Total Net Income (Sim)")
    _cim_amort       = _cim_col("Total Current Amortization")
    _cim_pct_amort   = _cim_col("% Net → Amort")
    _cim_risk        = _cim_col("Risk Label")

    # NIA sheet fixed column letters (C=3, D=4, E=5, F=6)
    _nia_net_col     = "C"   # Net Income + Amortization
    _nia_pct_col     = "D"   # % New Net Income to Amortization
    _nia_risk_col    = "E"   # Risk Label

    _CMP_NUM_FMT = '"₱"#,##0.00;"-₱"#,##0.00'
    _CMP_PCT_FMT = '0.0"%"'
    _CMP_DLT_FMT = '"+₱"#,##0.00;"-₱"#,##0.00;"-"'
    _CMP_DPT_FMT = '"+0.0%";"-0.0%";"-"'

    _data_start_cmp = 5

    for ri, _row in enumerate(merged_rows):
        row_num  = 5 + ri
        src_row  = 4 + ri   # data starts at row 4 in "Client Impact Merged"

        _meta    = _nia_row_meta[ri]   # pre-computed NIA values

        # Base row background
        _base_bg = _C_ROW_ODD if ri % 2 == 0 else _C_ROW_EVEN
        _cmp_rf  = PatternFill("solid", fgColor=_base_bg)

        _f_base  = Font(size=9, color="374151",  name="Calibri")
        _f_num   = Font(size=9, color=_C_NUM,    name="Calibri")
        _al_l    = Alignment(horizontal="left",   vertical="center")
        _al_c    = Alignment(horizontal="center", vertical="center")
        _al_r    = Alignment(horizontal="right",  vertical="center")

        def _plain(col_n, val, fmt=None, font=None, align=None, bg=None):
            _c = ws_cmp.cell(row_num, col_n, val)
            _c.fill      = PatternFill("solid", fgColor=(bg or _base_bg))
            _c.font      = font or _f_base
            _c.alignment = align or _al_l
            _c.border    = _grid_c
            if fmt:
                _c.number_format = fmt
            return _c

        # Col 1: Client ID
        if _cim_client_id:
            _plain(1, f"={_cim_sheet}!{_cim_client_id}{src_row}")
        else:
            _plain(1, str(_row.get("Client ID") or ""))

        # Col 2: Applicant
        if _cim_applicant:
            _plain(2, f"={_cim_sheet}!{_cim_applicant}{src_row}")
        else:
            _plain(2, str(_row.get("Applicant") or ""))

        # ── NET INCOME COMPARISON ─────────────────────────────────────
        # Col 3: [CIM] Total Net Income (Sim) — cross-sheet formula
        if _cim_sim_net:
            _plain(3, f"={_cim_sheet}!{_cim_sim_net}{src_row}",
                   fmt=_CMP_NUM_FMT, font=_f_num, align=_al_r)
        else:
            _plain(3, "", align=_al_r)

        # Col 4: [NIA] Net Income + Amortization — static value from pre-computed meta
        # Use red font when the value is negative so it is visually distinct.
        _c4_font = Font(size=9, color="C0392B" if _meta["nia_net"] < 0 else _C_NUM, name="Calibri")
        _plain(4, _meta["nia_net"], fmt=_CMP_NUM_FMT, font=_c4_font, align=_al_r)

        # Col 5: Total Current Amortization — cross-sheet formula from CIM for basis
        if _cim_amort:
            _plain(5, f"={_cim_sheet}!{_cim_amort}{src_row}",
                   fmt=_CMP_NUM_FMT, font=_f_num, align=_al_r)
        else:
            _plain(5, _meta["amort"], fmt=_CMP_NUM_FMT, font=_f_num, align=_al_r)

        # ── PERCENTAGE COMPARISON ─────────────────────────────────────
        # Col 6: [CIM] % Net → Amort — cross-sheet formula
        if _cim_pct_amort:
            _plain(6, f"={_cim_sheet}!{_cim_pct_amort}{src_row}",
                   fmt=_CMP_PCT_FMT, align=_al_c)
        else:
            _plain(6, "", align=_al_c)

        # Col 7: [NIA] % New Net Income to Amort — static value from pre-computed meta.
        # Sentinel -2.0 means zero amortization (N/A); -1.0 means negative income.
        if _meta["nia_pct"] == -2.0:
            _c7 = ws_cmp.cell(row_num, 7, "N/A (No Amortization)")
            _c7.fill      = PatternFill("solid", fgColor=_base_bg)
            _c7.font      = Font(italic=True, size=9, color=_C_META_FG, name="Calibri")
            _c7.alignment = _al_c
            _c7.border    = _grid_c
        elif _meta["nia_pct"] < 0:
            _c7 = ws_cmp.cell(row_num, 7, _NEGATIVE_INCOME_PCT_DISPLAY)
            _c7.fill      = PatternFill("solid", fgColor=_base_bg)
            _c7.font      = Font(italic=True, size=9, color=_C_HIGH_FG, name="Calibri")
            _c7.alignment = _al_c
            _c7.border    = _grid_c
        else:
            _plain(7, _meta["nia_pct"], fmt=_CMP_PCT_FMT, align=_al_c)

        # Col 8: Delta % (NIA pct − CIM pct) — both evaluated on same raw data,
        # different denominators, so the sign is meaningful.
        # When NIA pct is either sentinel, delta is not meaningful; show "—".
        if _meta["nia_pct"] < 0:
            _c8 = ws_cmp.cell(row_num, 8, "—")
            _c8.fill      = PatternFill("solid", fgColor=_C_DELTA_NEU)
            _c8.font      = Font(italic=True, size=9, color=_C_HIGH_FG, name="Calibri")
            _c8.alignment = _al_c
            _c8.border    = _grid_c
        else:
            try:
                _cim_pct_v = float(str(_row.get("% Net \u2192 Amort") or "0").replace("%",""))
            except Exception:
                _cim_pct_v = 0.0
            _dp     = _meta["nia_pct"] - _cim_pct_v
            _c8_bg  = _C_DELTA_POS if _dp > 0.005 else (_C_DELTA_NEG if _dp < -0.005 else _C_DELTA_NEU)
            _plain(8, _dp, fmt=_CMP_PCT_FMT,
                   font=Font(size=9, bold=True, color="374151", name="Calibri"),
                   align=_al_c, bg=_c8_bg)

        # ── RISK LABEL COMPARISON ─────────────────────────────────────
        # CIM risk label — from the merged_rows dict (written statically to CIM sheet)
        _cim_risk_val = str(_row.get("Risk Label") or "").upper()
        _nia_risk_val = _meta["nia_risk"]

        # Col 9: [CIM] Risk Label — cross-sheet formula to CIM sheet
        if _cim_risk:
            _plain(9, f"={_cim_sheet}!{_cim_risk}{src_row}", align=_al_c,
                   font=Font(size=9, bold=True,
                             color=_risk_colors.get(_cim_risk_val, (None, "374151", False))[1],
                             name="Calibri"))
        else:
            _plain(9, _cim_risk_val, align=_al_c)

        # Col 10: [NIA] Risk Label — static value (NIA computes its own label)
        _plain(10, _nia_risk_val, align=_al_c,
               font=Font(size=9, bold=True,
                         color=_risk_colors.get(_nia_risk_val, (None, "374151", False))[1],
                         name="Calibri"))

        # Col 11: Label Changed? — compare the two Python-computed risk labels
        _labels_differ = (_cim_risk_val != _nia_risk_val)
        _changed_text  = "YES \u26a0" if _labels_differ else "NO \u2713"
        _c11 = ws_cmp.cell(row_num, 11, _changed_text)
        _c11.alignment = _al_c
        _c11.border    = _grid_c
        _c11.font      = Font(size=9, bold=True, name="Calibri",
                              color=(_C_HIGH_FG if _labels_differ else "1E8449"))
        _c11.fill      = PatternFill("solid",
                                     fgColor=(_C_CHANGED if _labels_differ else _C_SAME))

        ws_cmp.row_dimensions[row_num].height = 18

    # ── TOTAL / AVERAGE summary rows ─────────────────────────────────
    _cmp_data_end  = 4 + len(merged_rows)
    _cmp_total_row = _cmp_data_end + 1
    _cmp_avg_row   = _cmp_data_end + 2

    _SUM_COLS_CMP = {3, 4, 5}   # numeric ₱ columns
    _PCT_COLS_CMP = {6, 7, 8}   # percentage columns

    def _write_cmp_summary(row_n, label, bg, lbl_f, num_f, brd):
        rf = PatternFill("solid", fgColor=bg)
        for ci2 in range(1, _cmp_total_cols + 1):
            c2 = ws_cmp.cell(row_n, ci2)
            c2.fill = rf; c2.border = brd
            cl = get_column_letter(ci2)
            if ci2 == 1:
                c2.value = label; c2.font = lbl_f
                c2.alignment = Alignment(horizontal="left", vertical="center")
            elif ci2 in _SUM_COLS_CMP:
                c2.value = (
                    f"=SUM({cl}{_data_start_cmp}:{cl}{_cmp_data_end})"
                    if label == "TOTAL"
                    else f"=AVERAGE({cl}{_data_start_cmp}:{cl}{_cmp_data_end})"
                )
                c2.number_format = _CMP_NUM_FMT
                c2.font = num_f
                c2.alignment = Alignment(horizontal="right", vertical="center")
            elif ci2 in _PCT_COLS_CMP:
                c2.value = (
                    "" if label == "TOTAL"
                    else f"=AVERAGE({cl}{_data_start_cmp}:{cl}{_cmp_data_end})"
                )
                c2.number_format = _CMP_PCT_FMT
                c2.font = num_f
                c2.alignment = Alignment(horizontal="center", vertical="center")
            else:
                c2.value = ""; c2.font = lbl_f
                c2.alignment = Alignment(horizontal="left", vertical="center")
        ws_cmp.row_dimensions[row_n].height = 18

    ws_cmp.freeze_panes = "A5"
    ws_cmp.auto_filter.ref = f"A4:{get_column_letter(_cmp_total_cols)}{4 + len(merged_rows)}"

    wb.save(out_path)
    matched_count = len(impact_rows) - sum(
        1 for r in impact_rows if _norm_name(r.get("Applicant", "")) not in matched_keys
    )
    messagebox.showinfo(
        "Merge Complete",
        f"Merged {len(paths)} file(s) into Client Impact table\n"
        f"Client Impact rows:  {len(impact_rows)}\n"
        f"Matched by name:     {matched_count}\n"
        f"Unmatched (appended): {unmatched}\n"
        f"Output columns:      {len(all_output_headers)}\n\n"
        f"Saved to:\n{out_path}",
        parent=self,
    )


def _sim_export_high_risk_clients_excel(self):
    """
    Export only HIGH-risk clients from the simulator "Client Impact" table to Excel.

    - HIGH risk is defined as % Net → Amort >= 71%.
    - Uses current simulator slider values (what-if).
    - Uses the current LU industry filter already applied to `_sim_recs`.
    - Uses the current client search filter from `_sim_client_search_var`.
    - Exports all columns identical to the Client Impact table, including Industry.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messagebox.showerror(
            "Missing Library",
            "openpyxl is not installed.\nRun:  pip install openpyxl",
            parent=self,
        )
        return

    recs_all = list(getattr(self, "_sim_recs", []) or [])
    if not recs_all:
        messagebox.showinfo(
            "No simulator data",
            "Run LU analysis and the simulator first, then try Export again.",
            parent=self,
        )
        return

    # Apply client name search filter (same logic as the on-screen table).
    _search_term = str(
        getattr(self, "_sim_client_search_var", None) and self._sim_client_search_var.get()
        or ""
    ).strip().lower()
    if _search_term:
        recs_all = [
            r for r in recs_all
            if _search_term in str((r or {}).get("client") or "").lower()
        ]

    # Build per-client simulated totals and risk labels (same rules as the table).
    rows = []
    for rec in recs_all:
        name = str((rec or {}).get("client") or "").strip() or "—"
        industry = str((rec or {}).get("industry") or "").strip() or "—"
        base_net = float((rec or {}).get("net_income") or 0.0)
        current_am = float((rec or {}).get("current_amort") or 0.0)

        base_by_name: dict[str, float] = {}
        base_total_exp = 0.0
        for exp in (rec or {}).get("expenses", []) or []:
            nm = str((exp or {}).get("name") or "").strip()
            if not nm:
                continue
            try:
                amt = float((exp or {}).get("total") or 0.0)
            except Exception:
                amt = 0.0
            if amt <= 0:
                continue
            base_by_name[nm] = base_by_name.get(nm, 0.0) + amt
            base_total_exp += amt

        extra_total = 0.0
        for nm, base_amt in base_by_name.items():
            extra, _sim = _sim_amount_for_expense(self, base_amt, nm)
            extra_total += extra

        sim_total_exp = base_total_exp + extra_total
        pct_inc = (
            extra_total / base_total_exp * 100.0
            if base_total_exp > 0
            else (0.0 if extra_total <= 0 else 100.0)
        )

        sim_net_income = base_net - extra_total
        if current_am == 0:
            pct_net_to_am = -2.0  # sentinel: no amortization — risk is N/A
        elif sim_net_income <= 0:
            pct_net_to_am = -1.0  # sentinel: negative income
        else:
            pct_net_to_am = (current_am / sim_net_income) * 100.0

        sim_risk = _sim_pct_net_to_amort_label(pct_net_to_am, getattr(self, "_sim_risk_ranges", None))
        risk_reasoning = _sim_build_risk_reasoning(sim_risk, pct_net_to_am)
        rows.append({
            "client_id": str((rec or {}).get("client_id") or "").strip(),
            "pn": str((rec or {}).get("pn") or "").strip(),
            "client": name,
            "residence_address": str((rec or {}).get("residence_address") or "").strip(),
            "office_address": str((rec or {}).get("office_address") or "").strip(),
            "industry": industry,
            "loan_status": str((rec or {}).get("loan_status") or "").strip(),
            "ao_name": str((rec or {}).get("ao_name") or "").strip(),
            "ci_bi_date": str((rec or {}).get("ci_bi_date") or "").strip(),
            "product_name": str((rec or {}).get("product_name") or "").strip(),
            "loan_balance": float((rec or {}).get("loan_balance") or 0.0),
            "principal_loan": float((rec or {}).get("principal_loan") or 0.0),
            "base_total_expenses": base_total_exp,
            "sim_total_expenses": sim_total_exp,
            "net_income": base_net,
            "sim_net_income": sim_net_income,
            "pct_increase": pct_inc,
            "sim_increase": extra_total,
            "current_amort": current_am,
            "pct_net_to_amort": pct_net_to_am,
            "sim_risk_label": sim_risk,
            "risk_reasoning": risk_reasoning,
        })

    # Filter to HIGH risk only (>= 71% Net → Amort).
    high_risk_rows = [r for r in rows if r["sim_risk_label"] == "HIGH" or "NEGATIVE" in str(r.get("sim_risk_label", "")).upper()]

    if not high_risk_rows:
        messagebox.showinfo(
            "No HIGH Risk Clients",
            (
                "No clients are currently classified as HIGH risk "
                "(% Net → Amort ≥ 71%) under the current simulator settings.\n\n"
                "Try increasing the inflation rates to see high-risk clients appear."
            ),
            parent=self,
        )
        return

    # Sort by highest % Net → Amort first (descending).
    high_risk_rows.sort(key=lambda r: -(r["pct_net_to_amort"] or 0.0))

    from tkinter import filedialog
    import getpass
    from datetime import datetime
    from pathlib import Path

    default_name = f"RiskSimulator_HIGH_Risk_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    path = filedialog.asksaveasfilename(
        parent=self,
        title="Save HIGH Risk Clients Excel",
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=default_name,
    )
    if not path:
        return

    wb = openpyxl.Workbook()
    ws_cfg = wb.active
    ws_cfg.title = "Settings"

    exported_by = str(getattr(self, "_current_username", "") or "").strip()
    if not exported_by:
        try:
            exported_by = (getpass.getuser() or "").strip()
        except Exception:
            exported_by = ""
    if not exported_by:
        exported_by = "Unknown user"

    selected_inds = sorted(getattr(self, "_sim_selected_industries", set()) or set(), key=str.lower)
    industry_note = "None (no industry filter)" if not selected_inds else " · ".join(selected_inds)
    search_note = _search_term if _search_term else "None"

    _sim_write_export_settings_sheet(
        ws_cfg,
        self_app=self,
        fname=Path(str(getattr(self, "_lu_filepath", "") or "—")).name,
        generated_at=datetime.now().strftime("%B %d, %Y  %H:%M"),
        exported_by=exported_by,
        export_scope_note=(
            "HIGH Risk Clients export — includes only clients with % Net → Amort ≥ 71% "
            f"under current simulator what-if values. Client search: {search_note}."
        ),
    )

    # ── Append formula/risk/expense/industry sections to the same sheet ──
    _used_rows_hr = ws_cfg.max_row
    _sim_write_settings_sheet(ws_cfg, self_app=self, row_offset=_used_rows_hr + 1)

    # HIGH Risk Clients sheet.
    ws = wb.create_sheet("HIGH Risk Clients")
    NUM_COLS = 22

    # ── Shared modern style palette (red/danger theme) ────────────────
    _C_HDR_BG  = "7B1A14"   # deep crimson header
    _C_HDR_FG  = "FFFFFF"
    _C_BAN_BG  = "A93226"   # banner crimson
    _C_BAN_FG  = "FFFFFF"
    _C_META_BG = "FDFEFE"
    _C_META_FG = "7B241C"
    _C_ROW_ODD  = "FFFFFF"
    _C_ROW_EVEN = "FEF9F9"   # very faint rose stripe
    _C_HIGH_BG  = "FDECEA"
    _C_HIGH_FG  = "C0392B"
    _C_BORDER   = "E8C8C5"
    _C_NUM      = "A93226"   # numbers in crimson accent

    _thin_r  = Side(border_style="thin",   color=_C_BORDER)
    _thick_r = Side(border_style="medium", color=_C_HDR_BG)
    _grid_r  = Border(left=_thin_r, right=_thin_r, top=_thin_r, bottom=_thin_r)
    _hdr_bot = Border(left=_thin_r, right=_thin_r, top=_thin_r, bottom=_thick_r)

    NUM_FMT = '"₱"#,##0.00'
    PCT_FMT = '0.0"%"'

    # Row 1 — Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NUM_COLS)
    _bann = ws.cell(1, 1, "🔴  HIGH RISK CLIENTS — % Net Income to Amortization ≥ 71%")
    _bann.fill      = PatternFill("solid", fgColor=_C_BAN_BG)
    _bann.font      = Font(bold=True, size=13, color=_C_BAN_FG, name="Calibri")
    _bann.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Row 2 — Meta info
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=NUM_COLS)
    summary_text = (
        f"Total HIGH risk clients: {len(high_risk_rows)}     "
        f"Exported: {datetime.now().strftime('%B %d, %Y  %H:%M')}     "
        f"Exported by: {exported_by}"
    )
    _m2 = ws.cell(2, 1, summary_text)
    _m2.fill      = PatternFill("solid", fgColor=_C_META_BG)
    _m2.font      = Font(italic=True, size=9, color=_C_META_FG, name="Calibri")
    _m2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 16

    # Row 3 — Column headers
    headers = [
        "Client ID", "PN", "Client", "Residence Address", "Office Address",
        "Industry", "Loan Status", "AO Name", "CI/BI Date", "Product Name",
        "Loan Balance", "Principal Loan",
        "Total Expenses (Base)", "Total Expenses (Sim)",
        "Total Net Income (Base)", "Total Net Income (Sim)",
        "% Increase", "Simulated Increase",
        "Total Current Amort", "% Net → Amort",
        "Risk Label", "Risk Reasoning",
    ]
    _hf  = Font(bold=True, size=9, color=_C_HDR_FG, name="Calibri")
    _ha  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(3, ci, h)
        cell.fill      = PatternFill("solid", fgColor=_C_HDR_BG)
        cell.font      = _hf
        cell.alignment = _ha
        cell.border    = _hdr_bot
    ws.row_dimensions[3].height = 32

    # Column widths
    col_widths = [13, 13, 24, 22, 22, 18, 14, 18, 15, 20, 15, 15, 18, 18, 18, 18, 11, 17, 17, 13, 12, 48]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Data rows (start row 4)
    for idx, r in enumerate(high_risk_rows):
        row_num  = 4 + idx
        row_bg   = _C_ROW_ODD if idx % 2 == 0 else _C_ROW_EVEN
        _rf      = PatternFill("solid", fgColor=row_bg)
        _base_f  = Font(size=9, color="374151", name="Calibri")
        _num_f   = Font(size=9, color=_C_NUM,   name="Calibri")
        _risk_f  = Font(bold=True, size=9, color=_C_HIGH_FG, name="Calibri")
        _al_l    = Alignment(horizontal="left",   vertical="center", wrap_text=False)
        _al_c    = Alignment(horizontal="center", vertical="center")
        _al_r    = Alignment(horizontal="right",  vertical="center")

        def _dc(col, value, font=_base_f, align=_al_l, fmt=None):
            cell = ws.cell(row_num, col, value)
            cell.fill = _rf; cell.font = font
            cell.border = _grid_r; cell.alignment = align
            if fmt: cell.number_format = fmt
            return cell

        _dc(1,  r["client_id"] or "—")
        _dc(2,  r["pn"] or "—")
        _dc(3,  r["client"])
        _dc(4,  r["residence_address"] or "—")
        _dc(5,  r["office_address"] or "—")
        _dc(6,  r["industry"])
        _dc(7,  r["loan_status"] or "—",       align=_al_c)
        _dc(8,  r["ao_name"] or "—")
        _dc(9,  r["ci_bi_date"] or "—")
        _dc(10, r["product_name"] or "—")
        _dc(11, r["loan_balance"],              font=_num_f, align=_al_r, fmt=NUM_FMT)
        _dc(12, r["principal_loan"],            font=_num_f, align=_al_r, fmt=NUM_FMT)
        _dc(13, r["base_total_expenses"],       font=_num_f, align=_al_r, fmt=NUM_FMT)
        _dc(14, r["sim_total_expenses"],        font=_num_f, align=_al_r, fmt=NUM_FMT)
        _dc(15, r["net_income"],                font=_num_f, align=_al_r, fmt=NUM_FMT)
        _dc(16, r["sim_net_income"],            font=_num_f, align=_al_r, fmt=NUM_FMT)
        _dc(17, r["pct_increase"],              font=_base_f, align=_al_c, fmt=PCT_FMT)
        _dc(18, r["sim_increase"],              font=_num_f, align=_al_r, fmt=NUM_FMT)
        _dc(19, r["current_amort"],             font=_num_f, align=_al_r, fmt=NUM_FMT)
        _dc(20, r["pct_net_to_amort"],          font=_base_f, align=_al_c, fmt=PCT_FMT)
        _dc(21, r["sim_risk_label"],            font=_risk_f, align=_al_c)
        _dc(22, r["risk_reasoning"],            align=Alignment(horizontal="left", vertical="center", wrap_text=True))
        ws.row_dimensions[row_num].height = 18

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(NUM_COLS)}{3 + len(high_risk_rows)}"

    wb.save(path)
    messagebox.showinfo(
        "Export Complete",
        f"HIGH Risk Clients Excel saved to:\n{path}\n\n"
        f"{len(high_risk_rows)} HIGH risk client(s) exported.",
        parent=self,
    )


def _sim_update_client_pagination_ui(self, *, total_rows: int):
    """Update Prev/Next buttons and page label for the simulator client table."""
    total = int(total_rows or 0)
    page = int(getattr(self, "_sim_client_page", 0) or 0)
    max_page = max(0, (total - 1) // SIM_CLIENT_PAGE_SIZE) if total else 0
    if page < 0:
        page = 0
    if page > max_page:
        page = max_page
    try:
        lbl = getattr(self, "_sim_client_page_lbl", None)
        if lbl is not None:
            lbl.config(text=f"Page {page + 1} of {max_page + 1 if total else 1}")
    except Exception:
        pass
    try:
        cnt = getattr(self, "_sim_client_count_lbl", None)
        if cnt is not None:
            start = page * SIM_CLIENT_PAGE_SIZE + 1 if total else 0
            end = min((page + 1) * SIM_CLIENT_PAGE_SIZE, total) if total else 0
            cnt.config(text=f"{start}–{end} of {total} client(s)")
    except Exception:
        pass
    try:
        prev_btn = getattr(self, "_sim_client_prev_btn", None)
        next_btn = getattr(self, "_sim_client_next_btn", None)
        if prev_btn is not None:
            prev_btn.configure(state=("normal" if total and page > 0 else "disabled"))
        if next_btn is not None:
            next_btn.configure(state=("normal" if total and page < max_page else "disabled"))
    except Exception:
        pass


def _sim_client_page_prev(self):
    self._sim_client_page = max(0, int(getattr(self, "_sim_client_page", 0) or 0) - 1)
    _sim_refresh_client_table(self)


def _sim_client_page_next(self):
    total = len(list(getattr(self, "_sim_recs", []) or []))
    max_page = max(0, (total - 1) // SIM_CLIENT_PAGE_SIZE) if total else 0
    self._sim_client_page = min(max_page, int(getattr(self, "_sim_client_page", 0) or 0) + 1)
    _sim_refresh_client_table(self)


# ══════════════════════════════════════════════════════════════════════
#  PIE CHART  (simulated expense mix — % of total)
# ══════════════════════════════════════════════════════════════════════

def _sim_draw_chart(self):
    holder = getattr(self, "_sim_chart_holder", None)
    if holder is None:
        return
    try:
        if not holder.winfo_exists():
            return
    except Exception:
        return

    for w in holder.winfo_children():
        w.destroy()

    expenses = [e for e in getattr(self, "_sim_expenses", []) if e["total"] > 0]
    if not expenses:
        tk.Label(
            holder,
            text="No numeric data\nto chart.",
            font=F(9),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
            justify="center",
        ).pack(pady=40)
        return

    expenses = expenses[:SIM_CHART_MAX_BARS]

    def _sim_amount(exp):
        pct = 0.0
        var = self._sim_sliders.get(exp["name"])
        if var:
            try:
                pct = float(var.get() or 0)
            except Exception:
                pass
        base = float(exp["total"] or 0)
        return max(0.0, base + base * (pct / 100.0))

    pairs = [(e["name"], _sim_amount(e)) for e in expenses]
    pairs.sort(key=lambda x: -x[1])
    names = [p[0] for p in pairs]
    vals = [p[1] for p in pairs]

    if sum(vals) <= 0:
        tk.Label(
            holder,
            text="No simulated amounts\nto chart.",
            font=F(9),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
            justify="center",
        ).pack(pady=40)
        return

    if len(pairs) > PIE_MAX_SLICES:
        top = pairs[: PIE_MAX_SLICES - 1]
        other_sum = sum(p[1] for p in pairs[PIE_MAX_SLICES - 1 :])
        names = [p[0] for p in top] + (["Other"] if other_sum > 0 else [])
        vals = [p[1] for p in top] + ([other_sum] if other_sum > 0 else [])

    if not _HAS_MPL:
        lines = [f"{n[:22]}{'…' if len(n) > 22 else ''}: {v/sum(vals)*100:.1f}%"
                 for n, v in zip(names, vals)]
        tk.Label(
            holder,
            text="matplotlib unavailable.\n\n" + "\n".join(lines[:12]),
            font=F(7),
            fg=_TXT_SOFT,
            bg=_CARD_WHITE,
            justify="left",
        ).pack(padx=6, pady=8)
        return

    def _short(n: str, w: int = 18) -> str:
        n = str(n or "").strip()
        return n if len(n) <= w else n[: w - 1] + "…"

    try:
        fig, ax = plt.subplots(figsize=(4.7, 4.9))
        fig.patch.set_facecolor(_CARD_WHITE)
        ax.set_facecolor(_CARD_WHITE)

        colors = [plt.cm.Pastel2(i % 8) for i in range(len(vals))]
        wedges, _texts, autotexts = ax.pie(
            vals,
            labels=None,
            colors=colors,
            startangle=90,
            counterclock=False,
            autopct=lambda p: f"{p:.1f}%" if p >= 4.5 else "",
            pctdistance=0.80,
            textprops={"fontsize": 8, "color": "#243B64", "fontweight": "bold"},
            wedgeprops={"width": 0.44, "linewidth": 1.0, "edgecolor": _CARD_WHITE},
        )
        for t in autotexts:
            t.set_fontsize(8)
        total_sim = sum(vals)
        ax.text(
            0, 0,
            f"Total\nP{total_sim:,.0f}",
            ha="center",
            va="center",
            fontsize=9,
            color="#365B8C",
            fontweight="bold",
        )
        ax.set_title("Share of Total (Simulated)", fontsize=9, color="#4A6FA5", pad=6)

        leg_labels = [_short(n, 22) for n in names]
        ax.legend(
            wedges,
            leg_labels,
            loc="upper center",
            bbox_to_anchor=(0.5, -0.07),
            ncol=2,
            fontsize=6.5,
            frameon=False,
        )
        fig.subplots_adjust(left=0.06, right=0.94, top=0.88, bottom=0.30)

        canvas = FigureCanvasTkAgg(fig, master=holder)
        widget = canvas.get_tk_widget()
        widget.config(width=370, height=370)
        widget.pack_propagate(False)
        widget.pack(fill="none", expand=False)
        plt.close(fig)
    except Exception:
        tk.Label(
            holder,
            text="Could not draw chart.",
            font=F(9),
            fg=_TXT_MUTED,
            bg=_CARD_WHITE,
        ).pack(pady=24)


# ══════════════════════════════════════════════════════════════════════
#  ATTACH
# ══════════════════════════════════════════════════════════════════════

def attach(cls):
    """
    Attach Risk Simulator methods to the app class.
    Call AFTER lu_tab_analysis.attach(cls).
    """
    cls._build_simulator_panel    = _build_simulator_panel
    cls._build_sim_summary_cards  = _build_sim_summary_cards
    cls._sim_show_placeholder     = _sim_show_placeholder
    cls._sim_open_expense_table_window = _sim_open_expense_table_window
    cls._sim_render_expense_table_rows = _sim_render_expense_table_rows
    cls._sim_populate             = _sim_populate
    cls._sim_build_expense_row    = _sim_build_expense_row
    cls._sim_on_slide             = _sim_on_slide
    cls._sim_apply_global         = _sim_apply_global
    cls._sim_reset                = _sim_reset
    cls._sim_refresh              = _sim_refresh
    cls._sim_draw_chart           = _sim_draw_chart
    cls._sim_export_client_impact_excel = _sim_export_client_impact_excel
    cls._sim_merge_excel_files = _sim_merge_excel_files
    cls._sim_export_high_risk_clients_excel = _sim_export_high_risk_clients_excel
    cls._sim_open_risk_ranges_dialog = _sim_open_risk_ranges_dialog
    cls._sim_on_client_impact_row_activated = _sim_on_client_impact_row_activated
    cls._sim_show_client_details  = _sim_show_client_details
    cls._sim_show_table_view      = _sim_show_table_view
    cls._sim_show_detail_view     = _sim_show_detail_view